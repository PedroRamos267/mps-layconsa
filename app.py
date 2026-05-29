import subprocess
import sqlite3
from datetime import date, timedelta
import os
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

import dash
from dash.exceptions import PreventUpdate
from dash import dcc, html, dash_table, Input, Output, State, callback_context
import dash_bootstrap_components as dbc
import pandas as pd

from db import (init_db, cargar_plan_desde_excel, cargar_bom_stock,
                get_ordenes_activas_en_dia, get_primera_fecha, get_mrp_dia,
                get_mrp_proyectado, get_cumplimiento_semanal, get_cumplimiento_detalle,
                get_semanas_plan, get_resumen_ops, get_avance_campana)

DB_PATH    = "mps_plasticos.db"
EXCEL_PLAN = "plan_produccion.xlsx"


def fmt_num(n):
    if n is None: return "0"
    try: return f"{int(n):,}"
    except: return "0"


def make_kpi(titulo, valor, subtitulo, color="light"):
    return dbc.Card(
        dbc.CardBody([
            html.P(titulo, className="text-muted mb-1", style={"fontSize": "12px"}),
            html.H4(valor, className=f"text-{color} mb-0", style={"fontWeight": "500"}),
            html.Small(subtitulo, className="text-muted"),
        ]),
        className="h-100",
        style={"borderRadius": "8px", "border": "0.5px solid #dee2e6"},
    )


def build_layout():
    fecha_inicial = get_primera_fecha(DB_PATH)
    try:
        fecha_label = date.fromisoformat(fecha_inicial).strftime("%d/%m/%Y")
    except Exception:
        fecha_label = fecha_inicial

    return dbc.Container([
        dcc.Store(id="store-fecha", data=fecha_inicial),
        dcc.Interval(id="auto-refresh", interval=60_000, n_intervals=0),

        # Header
        dbc.Row([
            dbc.Col([
                html.H5("MPS / MRP Plásticos – Layconsa", className="mb-0"),
                html.Small("Plan del día vs. producción real", className="text-muted"),
            ], width=6),
            dbc.Col([
                dcc.DatePickerSingle(
                    id="date-picker",
                    date=fecha_inicial,
                    display_format="DD/MM/YYYY",
                    first_day_of_week=1,
                    style={"float":"right","marginRight":"8px"},
                ),
                dbc.Button("📥 Exportar Excel", id="btn-exportar", color="success",
                           size="sm", outline=True, className="float-end me-2"),
            ], width=6),
        ], className="mb-3 pt-3 border-bottom pb-2"),

        html.Div(id="export-feedback", className="mb-2"),

        # Pestañas
        dbc.Tabs([

            # ── PESTAÑA MPS ──────────────────────────────────────────────────
            dbc.Tab(label="📋 MPS – Plan de Producción", tab_id="tab-mps", children=[
                html.Div(className="mt-3", children=[
                    dbc.Row(id="kpi-row", className="mb-3 g-2"),
                    dbc.Row(id="kpi-semanal", className="mb-2 g-2"),
                    dbc.Row([
                        dbc.Col([
                            dcc.Dropdown(id="fil-proceso", placeholder="Proceso",
                                         multi=True, style={"fontSize": "13px"}),
                        ], width=3),
                        dbc.Col([
                            dcc.Dropdown(id="fil-maquina", placeholder="Máquina",
                                         multi=True, style={"fontSize": "13px"}),
                        ], width=3),
                        dbc.Col([
                            dcc.Dropdown(id="fil-estado",
                                options=[
                                    {"label": "✅ Completado", "value": "ok"},
                                    {"label": "🟡 Parcial",    "value": "warn"},
                                    {"label": "🔴 Atrasado",   "value": "bad"},
                                    {"label": "⬜ Pendiente",  "value": "pend"},
                                ],
                                placeholder="Estado", multi=True,
                                style={"fontSize": "13px"}),
                        ], width=3),
                        dbc.Col([
                            dbc.Button("＋ Registrar avance", id="btn-abrir-modal",
                                       color="primary", size="sm", className="float-end"),
                        ], width=3),
                    ], className="mb-2 g-2"),
                    html.Div(id="tabla-container"),
                ]),
            ]),

            # ── PESTAÑA MRP ──────────────────────────────────────────────────
            dbc.Tab(label="🔩 MRP – Requerimiento de Materiales", tab_id="tab-mrp", children=[
                html.Div(className="mt-3", children=[
                    dbc.Tabs([
                        # Sub-pestaña: Resumen del día
                        dbc.Tab(label="📊 Resumen del día", tab_id="mrp-dia", children=[
                            html.Div(className="mt-3", children=[
                                dbc.Row(id="mrp-kpi-row", className="mb-3 g-2"),
                                dbc.Row([
                                    dbc.Col([
                                        dcc.Dropdown(id="mrp-fil-proceso", placeholder="Tipo de material",
                                                     multi=True, style={"fontSize": "13px"}),
                                    ], width=4),
                                    dbc.Col([
                                        dcc.Dropdown(id="mrp-fil-alerta",
                                            options=[
                                                {"label": "🔴 Falta stock", "value": "alerta"},
                                                {"label": "✅ Cubierto",    "value": "ok"},
                                            ],
                                            placeholder="Filtrar por cobertura",
                                            style={"fontSize": "13px"}),
                                    ], width=4),
                                    dbc.Col([
                                        dbc.Button("📥 Exportar MRP Excel", id="btn-exportar-mrp",
                                                   color="warning", size="sm", outline=True,
                                                   className="float-end"),
                                    ], width=4),
                                ], className="mb-2 g-2"),
                                html.Div(id="mrp-feedback", className="mb-2"),
                                html.Div(id="mrp-tabla-container"),
                            ]),
                        ]),
                        # Sub-pestaña: Proyección Comprados
                        dbc.Tab(label="⚠️ Proyección – Comprados", tab_id="mrp-proy-comp", children=[
                            html.Div(className="mt-3", children=[
                                dbc.Row(id="proy-kpi-row", className="mb-3 g-2"),
                                dbc.Row([
                                    dbc.Col([
                                        dcc.Dropdown(id="proy-fil-tipo",
                                                     placeholder="Tipo de material",
                                                     multi=True, style={"fontSize": "13px"}),
                                    ], width=3),
                                    dbc.Col([
                                        dcc.Dropdown(id="proy-fil-semaforo",
                                            options=[
                                                {"label": "🔴 Quiebre < 7 días",   "value": "rojo"},
                                                {"label": "🟡 Quiebre < 15 días",  "value": "amarillo"},
                                                {"label": "✅ Cobertura > 30 días", "value": "verde"},
                                            ],
                                            placeholder="Filtrar por semáforo",
                                            multi=True,
                                            style={"fontSize": "13px"}),
                                    ], width=4),
                                    dbc.Col([
                                        dbc.Button("📥 Exportar", id="btn-export-proy",
                                                   color="warning", size="sm", outline=True,
                                                   className="float-end me-2"),
                                        dbc.Button("🔄 Recalcular", id="btn-recalc-proy",
                                                   color="secondary", size="sm", outline=True,
                                                   className="float-end me-2"),
                                    ], width=5),
                                ], className="mb-2 g-2"),
                                html.Div(id="proy-feedback", className="mb-2"),
                                html.Div(id="proy-tabla-container"),
                            ]),
                        ]),
                        # Sub-pestaña: Proyección Semiterminados
                        dbc.Tab(label="🔧 Proyección – Semiterminados", tab_id="mrp-proy-semi", children=[
                            html.Div(className="mt-3", children=[
                                dbc.Row(id="semi-kpi-row", className="mb-3 g-2"),
                                dbc.Row([
                                    dbc.Col([
                                        dcc.Dropdown(id="semi-fil-tipo",
                                                     placeholder="Tipo de material",
                                                     multi=True, style={"fontSize": "13px"}),
                                    ], width=3),
                                    dbc.Col([
                                        dcc.Dropdown(id="semi-fil-semaforo",
                                            options=[
                                                {"label": "🔴 Quiebre < 7 días",   "value": "rojo"},
                                                {"label": "🟡 Quiebre < 15 días",  "value": "amarillo"},
                                                {"label": "✅ Cobertura > 30 días", "value": "verde"},
                                            ],
                                            placeholder="Filtrar por semáforo",
                                            multi=True,
                                            style={"fontSize": "13px"}),
                                    ], width=4),
                                    dbc.Col([
                                        dbc.Button("📥 Exportar", id="btn-export-semi",
                                                   color="warning", size="sm", outline=True,
                                                   className="float-end me-2"),
                                        dbc.Button("🔄 Recalcular", id="btn-recalc-semi",
                                                   color="secondary", size="sm", outline=True,
                                                   className="float-end me-2"),
                                    ], width=5),
                                ], className="mb-2 g-2"),
                                html.Div(id="semi-feedback", className="mb-2"),
                                html.Div(id="semi-tabla-container"),
                            ]),
                        ]),
                    ], id="mrp-subtabs", active_tab="mrp-dia"),
                ]),
            ]),

            # ── PESTAÑA RESUMEN OPS ─────────────────────────────────────────────
            dbc.Tab(label="📋 Resumen OPs", tab_id="tab-resumen", children=[
                html.Div(className="mt-3", children=[
                    dbc.Row([
                        dbc.Col([
                            dcc.Dropdown(id="res-fil-proceso", placeholder="Proceso",
                                         multi=True, style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dcc.Dropdown(id="res-fil-linea", placeholder="Línea",
                                         multi=True, style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dcc.Dropdown(id="res-fil-estado",
                                options=[
                                    {"label":"✅ Completado",  "value":"Completado"},
                                    {"label":"🔵 En curso",    "value":"En curso"},
                                    {"label":"🟡 Parcial",     "value":"Parcial"},
                                    {"label":"⬜ Pendiente",   "value":"Pendiente"},
                                ],
                                placeholder="Estado", multi=True,
                                style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dbc.Input(id="res-fil-op", placeholder="Buscar OP...",
                                      type="text", size="sm",
                                      style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dbc.Button("🔄 Actualizar", id="btn-res-refresh",
                                       color="secondary", size="sm", outline=True,
                                       className="float-end"),
                        ], width=4),
                    ], className="mb-3 g-2"),
                    dbc.Row(id="res-kpi-row", className="mb-3 g-2"),
                    html.Div(id="res-tabla"),
                ]),
            ]),

            # ── PESTAÑA AVANCE CAMPAÑA ──────────────────────────────────────────
            dbc.Tab(label="🏭 Avance campaña", tab_id="tab-campana", children=[
                html.Div(className="mt-3", children=[
                    dbc.Row([
                        dbc.Col([
                            dcc.Dropdown(id="camp-fil-proceso", placeholder="Proceso",
                                         multi=True, style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dcc.Dropdown(id="camp-fil-maquina", placeholder="Máquina",
                                         multi=True, style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dcc.Dropdown(id="camp-fil-linea", placeholder="Línea",
                                         multi=True, style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dcc.Dropdown(id="camp-fil-estado",
                                options=[
                                    {"label":"🔴 Crítico",      "value":"Critico"},
                                    {"label":"🟠 Retraso",      "value":"Retraso"},
                                    {"label":"🟡 Leve retraso", "value":"Leve retraso"},
                                    {"label":"✅ Al día",       "value":"Al dia"},
                                ],
                                placeholder="Estado", multi=True,
                                style={"fontSize":"13px"}),
                        ], width=3),
                        dbc.Col([
                            dbc.Button("🔄 Actualizar", id="btn-camp-refresh",
                                       color="secondary", size="sm", outline=True,
                                       className="float-end"),
                        ], width=3),
                    ], className="mb-3 g-2"),
                    dbc.Row(id="camp-kpi-row", className="mb-3 g-2"),
                    html.Div(id="camp-tabla"),
                ]),
            ]),

            # ── PESTAÑA GANTT ───────────────────────────────────────────────────
            dbc.Tab(label="📅 Gantt", tab_id="tab-gantt", children=[
                html.Div(className="mt-3", children=[
                    dbc.Row([
                        dbc.Col([
                            dcc.Dropdown(id="gantt-fil-proceso", placeholder="Proceso",
                                         style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dcc.Dropdown(id="gantt-fil-mes",
                                         placeholder="Mes",
                                         style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dcc.Dropdown(id="gantt-fil-maquina",
                                         placeholder="Máquina", multi=True,
                                         style={"fontSize":"13px"}),
                        ], width=2),
                        dbc.Col([
                            dcc.Dropdown(id="gantt-fil-linea",
                                         placeholder="Línea", multi=True,
                                         style={"fontSize":"13px"}),
                        ], width=3),
                        dbc.Col([
                            dbc.Button("🔄 Cargar", id="btn-gantt-load",
                                       color="primary", size="sm",
                                       className="float-end"),
                        ], width=3),
                    ], className="mb-3 g-2"),
                    html.Div(id="gantt-container"),
                ]),
            ]),

            # ── PESTAÑA CUMPLIMIENTO ────────────────────────────────────────────
            dbc.Tab(label="📈 Cumplimiento del programa", tab_id="tab-cum", children=[
                html.Div(className="mt-3", children=[
                    dbc.Tabs([
                        dbc.Tab(label="📅 Semanal", tab_id="cum-sub-semanal", children=[
                            html.Div(className="mt-3", children=[
                                dbc.Row([
                                    dbc.Col([
                                        dcc.Dropdown(id="cum-fil-semana", placeholder="Semana",
                                                     multi=True, style={"fontSize":"13px"}),
                                    ], width=3),
                                    dbc.Col([
                                        dcc.Dropdown(id="cum-fil-proceso", placeholder="Proceso",
                                                     multi=True, style={"fontSize":"13px"}),
                                    ], width=3),
                                    dbc.Col([
                                        dbc.Button("🔄 Actualizar", id="btn-cum-refresh",
                                                   color="secondary", size="sm", outline=True,
                                                   className="me-2"),
                                        dbc.Button("⬇️ Descargar semana", id="btn-cum-descargar",
                                                   color="success", size="sm", outline=True),
                                        dcc.Download(id="cum-download"),
                                    ], width=6, className="d-flex align-items-center justify-content-end gap-2"),
                                ], className="mb-3 g-2"),
                                dbc.Row(id="cum-kpi-row", className="mb-3 g-2"),
                                dbc.Row([
                                    dbc.Col(html.Div(id="cum-panel-dias"),  width=7),
                                    dbc.Col(html.Div(id="cum-panel-proc"),  width=5),
                                ], className="mb-3 g-2"),
                                html.Div(id="cum-tabla"),
                            ]),
                        ]),
                        dbc.Tab(label="📊 Resumen Campaña", tab_id="cum-sub-campana", children=[
                            html.Div(className="mt-3", children=[
                                dbc.Row([
                                    dbc.Col([dcc.Dropdown(id="cumres-fil-semana", placeholder="Semana", multi=True, style={"fontSize":"13px"})], width=2),
                                    dbc.Col([dcc.Dropdown(id="cumres-fil-mes", placeholder="Mes", style={"fontSize":"13px"})], width=2),
                                    dbc.Col([dcc.Dropdown(id="cumres-fil-linea", placeholder="Línea", multi=True, style={"fontSize":"13px"})], width=2),
                                    dbc.Col([dcc.Dropdown(id="cumres-fil-sub", placeholder="Subcomponente", multi=True, style={"fontSize":"13px"})], width=2),
                                    dbc.Col([dbc.Button("🔄 Actualizar", id="btn-cumres-refresh", color="secondary", size="sm", outline=True, className="float-end")], width=4),
                                ], className="mb-3 g-2"),
                                dbc.Row(id="cumres-kpi-row", className="mb-3 g-2"),
                                html.Div(id="cumres-graficos", className="mb-4"),
                                html.Div(id="cumres-tabla"),
                            ]),
                        ]),
                    ], id="cum-subtabs", active_tab="cum-sub-semanal"),
                ]),
            ]),

        ], id="tabs", active_tab="tab-mps"),

        # Modal registro avance
        dbc.Modal([
            dbc.ModalHeader(dbc.ModalTitle("Registrar avance real")),
            dbc.ModalBody([
                dbc.Label("Orden de producción", size="sm"),
                dcc.Dropdown(id="m-op", style={"fontSize": "13px"}),
                dbc.Row([
                    dbc.Col([
                        dbc.Label("Turno mañana", size="sm", className="mt-3"),
                        dbc.Input(id="m-qty-man", type="number", min=0, placeholder="0"),
                    ], width=6),
                    dbc.Col([
                        dbc.Label("Turno tarde", size="sm", className="mt-3"),
                        dbc.Input(id="m-qty-tar", type="number", min=0, placeholder="0"),
                    ], width=6),
                ]),
                dbc.Label("Total del día", size="sm", className="mt-3"),
                dbc.Input(id="m-qty", type="number", min=0, placeholder="0",
                          disabled=True, style={"fontWeight":"700","background":"#e9ecef"}),
                dbc.Label("Operario (opcional)", size="sm", className="mt-3"),
                dbc.Input(id="m-usuario", type="text", placeholder="Ej: Juan Pérez"),
                html.Div(id="m-feedback", className="mt-2"),
            ]),
            dbc.ModalFooter([
                dbc.Button("Cancelar", id="btn-cancelar", color="light", size="sm"),
                dbc.Button("Guardar",  id="btn-guardar",  color="primary", size="sm"),
            ]),
        ], id="modal", is_open=False),

    ], fluid=True, style={"maxWidth": "1500px", "fontFamily": "Arial, sans-serif"})


app = dash.Dash(__name__, external_stylesheets=[dbc.themes.BOOTSTRAP])
app.title = "MPS/MRP Plásticos – Layconsa"
app.layout = build_layout


# ── Callbacks MPS ─────────────────────────────────────────────────────────────

@app.callback(
    Output("store-fecha", "data"),
    Input("date-picker", "date"),
    prevent_initial_call=False,
)
def cambiar_dia(fecha_str):
    if not fecha_str:
        return str(date.today())
    return str(date.fromisoformat(str(fecha_str)[:10]))



@app.callback(
    Output("m-qty", "value"),
    Input("m-qty-man", "value"),
    Input("m-qty-tar", "value"),
)
def sumar_turnos(man, tar):
    return (man or 0) + (tar or 0)

@app.callback(
    Output("kpi-row", "children"),
    Output("tabla-container", "children"),
    Output("fil-proceso", "options"),
    Output("fil-maquina", "options"),
    Output("m-op", "options"),
    Input("store-fecha", "data"),
    Input("fil-proceso", "value"),
    Input("fil-maquina", "value"),
    Input("fil-estado", "value"),
    Input("auto-refresh", "n_intervals"),
)
def actualizar_mps(fecha_str, procesos, maquinas, estados, _):
    fecha = date.fromisoformat(fecha_str)
    fecha_label = fecha.strftime("%d/%m/%Y")
    df = get_ordenes_activas_en_dia(DB_PATH, fecha)

    if df.empty:
        return (
            [dbc.Col(make_kpi("Sin datos", "—", "No hay órdenes activas"), width=4)],
            dbc.Alert([html.Strong(fecha_label), " — Sin órdenes. Usá ‹ › para navegar."],
                      color="info", className="mt-2"),
            [], [], [],
        )

    proc_opts = [{"label": p, "value": p} for p in sorted(df["Proceso"].unique())]
    maq_opts  = [{"label": m, "value": m} for m in sorted(df["Maquina"].unique())]
    # Excluir CERRADAS/COMPLETADAS del dropdown de registro
    df_reg = df[df.get("Estado_OP", pd.Series(["ABIERTO"]*len(df))).str.upper().ne("CERRADO")] if "Estado_OP" in df.columns else df
    op_opts = [
        {"label": f"{int(r.BELNR_ID)} – {r.Maquina} – {r.Descripcion[:30]}",
         "value": r.BELNR_ID}
        for _, r in df_reg.iterrows()
    ]

    df["pct_dia"]  = (df["Real_Dia"] / df["Plan_Dia"] * 100).round(1).fillna(0).clip(upper=100)
    df["pct_acum"] = (df["Acumulado_Real"] / df["Planificado"] * 100).round(1).fillna(0)

    def clasif(p, estado_op="ABIERTO"):
        if str(estado_op).upper() == "CERRADO": return "completado"
        if p == 0:   return "pend"
        if p >= 100: return "ok"
        if p >= 75:  return "warn"
        return "bad"

    # Aplicar clasif con Estado_OP
    if "Estado_OP" in df.columns:
        df["Estado"] = df.apply(lambda r: clasif(r["pct_dia"], r.get("Estado_OP","ABIERTO")), axis=1)
    else:
        df["Estado"] = df["pct_dia"].apply(clasif)

    df["Estado_label"] = df["Estado"].map({
        "ok":         "✅ Completado",
        "completado": "✅ Completado",
        "warn":       "🟡 Parcial",
        "bad":        "🔴 Atrasado",
        "pend":       "⬜ Pendiente",
    })
    df["Inicio"] = df["Fecha_Inicio"].str[5:].str.replace("-","/") + " " + df["Hora_Inicio"]
    df["Fin"]    = df["Fecha_Fin"].str[5:].str.replace("-","/")    + " " + df["Hora_Fin"]
    df["Turnos"] = df["Horas_Prog"].apply(lambda h: "2 turnos" if h == 24 else "1 turno")

    if procesos: df = df[df["Proceso"].isin(procesos)]
    if maquinas: df = df[df["Maquina"].isin(maquinas)]
    if estados:  df = df[df["Estado"].isin(estados)]

    total = len(df)
    ok    = (df["Estado"] == "ok").sum()
    bad   = (df["Estado"] == "bad").sum()
    plan_tot = df["Plan_Dia"].sum()
    real_tot = df["Real_Dia"].sum()
    gp = int(real_tot / plan_tot * 100) if plan_tot > 0 else 0

    kpis = dbc.Row([
        dbc.Col(make_kpi("Órdenes activas", str(total), fecha_label, "info"), width=3),
        dbc.Col(make_kpi("Completadas", str(ok), f"de {total}", "success"), width=3),
        dbc.Col(make_kpi("Atrasadas", str(bad), "requieren atención",
                         "danger" if bad else "secondary"), width=3),
        dbc.Col(make_kpi("Avance del día", f"{gp}%",
                         f"{fmt_num(real_tot)} / {fmt_num(plan_tot)} und.",
                         "success" if gp >= 90 else "warning" if gp >= 70 else "danger"), width=3),
    ], className="g-2")

    COLOR_MAP = {"ok":"#c6efce","completado":"#c6efce","warn":"#ffeb9c","bad":"#ffc7ce","pend":"#f8f9fa"}
    tabla = dash_table.DataTable(
        id="tabla-mps",
        columns=[
            {"name":"O. Prod.",    "id":"BELNR_ID"},
            {"name":"Proceso",     "id":"Proceso"},
            {"name":"Línea",       "id":"Linea"},
            {"name":"Máquina",     "id":"Maquina"},
            {"name":"Turnos",      "id":"Turnos"},
            {"name":"Sec",         "id":"Sec"},
            {"name":"Código",      "id":"ItemCode"},
            {"name":"Descripción", "id":"Descripcion"},
            {"name":"Inicio",      "id":"Inicio"},
            {"name":"Fin",         "id":"Fin"},
            {"name":"Dur.(hs)",    "id":"Duracion_Horas", "type":"numeric","format":{"specifier":".1f"}},
            {"name":"Planificado", "id":"Planificado",    "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Plan día",    "id":"Plan_Dia",       "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Real día",    "id":"Real_Dia",       "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"% día",       "id":"pct_dia",        "type":"numeric","format":{"specifier":".1f"}},
            {"name":"Acumulado",   "id":"Acumulado_Real", "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Pendiente",   "id":"Pendiente",      "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"% acum.",     "id":"pct_acum",       "type":"numeric","format":{"specifier":".1f"}},
            {"name":"Estado",      "id":"Estado_label"},
        ],
        data=df.to_dict("records"),
        page_size=50, sort_action="native", filter_action="native",
        style_table={"overflowX":"auto"},
        style_header={"backgroundColor":"#1f3864","color":"white",
                      "fontWeight":"500","fontSize":"11px","border":"0.5px solid #dee2e6"},
        style_cell={"fontSize":"11px","padding":"6px 8px",
                    "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_data_conditional=[
            {"if":{"filter_query": '{Estado} = "' + e + '"'},
             "backgroundColor":c} for e,c in COLOR_MAP.items()
        ],
        style_cell_conditional=[
            {"if":{"column_id":"Descripcion"},"textAlign":"left","minWidth":"180px"},
            {"if":{"column_id":"Pendiente"},  "color":"#c0392b","fontWeight":"500"},
            {"if":{"column_id":"Acumulado_Real"},"color":"#1a7a4a","fontWeight":"500"},
        ],
    )
    return kpis, tabla, proc_opts, maq_opts, op_opts


# ── Callbacks MRP ─────────────────────────────────────────────────────────────


@app.callback(
    Output("kpi-semanal", "children"),
    Input("store-fecha", "data"),
    Input("auto-refresh", "n_intervals"),
)
def actualizar_semanal(fecha_str, _):
    try:
        fecha = date.fromisoformat(fecha_str)
        r = get_cumplimiento_semanal(DB_PATH, fecha)
        if not r["detalle"]:
            return []

        pct = r["pct_semana"]
        color = "success" if pct >= 90 else "warning" if pct >= 70 else "danger"

        # KPI resumen semana
        cards = [
            dbc.Col(dbc.Card(dbc.CardBody([
                html.P(f"Semana {r['semana']}", className="text-muted mb-1",
                       style={"fontSize":"11px"}),
                html.H5(f"{pct}%", className=f"text-{color} mb-0",
                        style={"fontWeight":"500"}),
                html.Small(f"{fmt_num(r['real_total'])} / {fmt_num(r['plan_total'])} und.",
                           className="text-muted"),
            ]), style={"borderRadius":"8px","border":"0.5px solid #dee2e6"}), width=2),
        ]

        # Mini barras por día
        for d in r["detalle"]:
            p = d["pct"]
            bc = "#1a7a4a" if p >= 90 else "#856404" if p >= 70 else "#842029"
            cards.append(dbc.Col(dbc.Card(dbc.CardBody([
                html.P(d["fecha"], className="text-muted mb-1",
                       style={"fontSize":"10px"}),
                html.Div(style={"background":"#e9ecef","borderRadius":"3px",
                                "height":"6px","marginBottom":"4px"},
                         children=[html.Div(style={
                             "width":f"{min(p,100)}%","height":"100%",
                             "background":bc,"borderRadius":"3px"})]),
                html.Small(f"{p:.0f}%", style={"fontSize":"11px","color":bc,
                                                "fontWeight":"500"}),
            ]), style={"borderRadius":"8px","border":"0.5px solid #dee2e6",
                       "padding":"6px"}), width=2))

        return cards
    except Exception:
        return []

@app.callback(
    Output("mrp-kpi-row", "children"),
    Output("mrp-tabla-container", "children"),
    Output("mrp-fil-proceso", "options"),
    Input("store-fecha", "data"),
    Input("mrp-fil-proceso", "value"),
    Input("mrp-fil-alerta", "value"),
    Input("auto-refresh", "n_intervals"),
    Input("tabs", "active_tab"),
)
def actualizar_mrp(fecha_str, procesos, alerta_fil, _, tab):
    # Solo recalcular cuando la pestaña MRP está activa
    if tab and "mrp" not in str(tab):
        raise PreventUpdate
    fecha = date.fromisoformat(fecha_str)
    fecha_label = fecha.strftime("%d/%m/%Y")

    try:
        df = get_mrp_dia(DB_PATH, fecha)
    except Exception as e:
        return (
            [],
            dbc.Alert(f"Error al calcular MRP: {str(e)}", color="danger"),
            [],
        )

    if df.empty:
        return (
            [],
            dbc.Alert(f"{fecha_label} — Sin plan activo o BOM no cargado. "
                      "Verificá que el Excel tenga hojas BOM y Stock.",
                      color="info"),
            [],
        )

    proc_opts = [{"label": p, "value": p} for p in sorted(df["Tipo_Item"].unique())]

    if procesos: df = df[df["Tipo_Item"].isin(procesos)]
    if alerta_fil == "alerta": df = df[df["Alerta"] == True]
    if alerta_fil == "ok":     df = df[df["Alerta"] == False]

    total_comp  = len(df)
    con_alerta  = df["Alerta"].sum()
    req_bruto   = df["Req_Bruto"].sum()
    req_neto    = df["Req_Neto"].sum()

    kpis = dbc.Row([
        dbc.Col(make_kpi("Componentes", str(total_comp), fecha_label, "info"), width=3),
        dbc.Col(make_kpi("Con alerta", str(int(con_alerta)),
                         "stock insuficiente", "danger" if con_alerta else "success"), width=3),
        dbc.Col(make_kpi("Req. bruto total", fmt_num(req_bruto), "unidades necesarias", "info"), width=3),
        dbc.Col(make_kpi("Req. neto total", fmt_num(req_neto),
                         "a comprar/producir",
                         "danger" if req_neto > 0 else "success"), width=3),
    ], className="g-2")

    df["Alerta_label"] = df["Alerta"].apply(lambda x: "🔴 Falta stock" if x else "✅ Cubierto")

    tabla = dash_table.DataTable(
        id="tabla-mrp",
        columns=[
            {"name":"Tipo",          "id":"Tipo_Item"},
            {"name":"Categoría",     "id":"Tipo_Material2"},
            {"name":"Niv.","id":"Nivel"},
            {"name":"Código",        "id":"Codigo_Comp"},
            {"name":"Descripción",   "id":"Desc_Comp"},
            {"name":"Req. bruto",    "id":"Req_Bruto",    "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"ALMA002",       "id":"Stock_ALMA002","type":"numeric","format":{"specifier":",.0f"}},
            {"name":"ALMA089",       "id":"Stock_ALMA089","type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Stock prod.",   "id":"Stock_Prod",   "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Stock total",   "id":"Stock_Total",  "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Req. neto",     "id":"Req_Neto",     "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Diferencia",    "id":"Diferencia",   "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Cobertura",     "id":"Alerta_label"},
        ],
        data=df.to_dict("records"),
        page_size=50, sort_action="native", filter_action="native",
        style_table={"overflowX":"auto"},
        style_header={"backgroundColor":"#1f3864","color":"white",
                      "fontWeight":"500","fontSize":"11px","border":"0.5px solid #dee2e6"},
        style_cell={"fontSize":"11px","padding":"6px 8px",
                    "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_data_conditional=[
            {"if":{"filter_query":'{Alerta_label} = "🔴 Falta stock"'},
             "backgroundColor":"#ffc7ce"},
            {"if":{"filter_query":'{Alerta_label} = "✅ Cubierto"'},
             "backgroundColor":"#c6efce"},
        ],
        style_cell_conditional=[
            {"if":{"column_id":"Desc_Comp"},   "textAlign":"left","minWidth":"200px"},
            {"if":{"column_id":"Tipo_Item"},   "textAlign":"left","minWidth":"160px"},
            {"if":{"column_id":"Tipo_Material2"},"textAlign":"left"},
            {"if":{"column_id":"Req_Neto"},    "fontWeight":"500","color":"#c0392b"},
            {"if":{"column_id":"Diferencia"},  "fontWeight":"700","color":"#c0392b","background":"#FCEBEB"},
            {"if":{"column_id":"Stock_Total"}, "fontWeight":"500","color":"#1a7a4a"},
        ],
    )
    return kpis, tabla, proc_opts




@app.callback(
    Output("proy-kpi-row", "children"),
    Output("proy-tabla-container", "children"),
    Output("proy-fil-tipo", "options"),
    Input("btn-recalc-proy", "n_clicks"),
    Input("proy-fil-tipo", "value"),
    Input("proy-fil-semaforo", "value"),
)
def actualizar_proyeccion(_, tipos, semaforos):
    try:
        df = get_mrp_proyectado(DB_PATH)
    except Exception as e:
        return [], dbc.Alert(f"Error: {str(e)}", color="danger"), []

    if df.empty:
        return [], dbc.Alert("Sin datos. Verificá BOM y Stock en el Excel.", color="info"), []

    # Solo comprados: excluir todos los 231xxx sin excepción
    df = df[~df["Codigo_Comp"].str.startswith("231")].copy()

    tipo_opts = [{"label": t, "value": t} for t in sorted(df["Tipo_Material2"].unique())]

    if tipos:    df = df[df["Tipo_Material2"].isin(tipos)]
    if semaforos: df = df[df["Semaforo"].isin(semaforos)]

    total     = len(df)
    n_rojo    = (df["Semaforo"] == "rojo").sum()
    n_amarillo= (df["Semaforo"] == "amarillo").sum()
    n_verde   = (df["Semaforo"] == "verde").sum()

    kpis = dbc.Row([
        dbc.Col(make_kpi("Componentes", str(total), "en el plan"), width=3),
        dbc.Col(make_kpi("🔴 Quiebre ≤ 3 días", str(int(n_rojo)),
                         "acción inmediata", "danger" if n_rojo else "secondary"), width=3),
        dbc.Col(make_kpi("🟡 Quiebre ≤ 7 días", str(int(n_amarillo)),
                         "planificar compra", "warning" if n_amarillo else "secondary"), width=3),
        dbc.Col(make_kpi("✅ Sin riesgo", str(int(n_verde)),
                         "cobertura > 7 días", "success"), width=3),
    ], className="g-2")

    # Columnas fijas + columnas de días
    cols_fijas = [
        {"name": "Semáforo",         "id": "Semaforo_icon"},
        {"name": "Tipo Material",     "id": "Tipo_Material2"},
        {"name": "Código",            "id": "Codigo_Comp"},
        {"name": "Descripción",       "id": "Desc_Comp"},
        {"name": "Stock inicial",     "id": "Stock_Inicial",       "type":"numeric","format":{"specifier":",.0f"}},
        {"name": "Consumo total",     "id": "Consumo_Total",       "type":"numeric","format":{"specifier":",.0f"}},
        {"name": "Diferencia",        "id": "Diferencia",          "type":"numeric","format":{"specifier":",.0f"}},
        {"name": "Cons./día prom.",   "id": "Consumo_Diario_Prom", "type":"numeric","format":{"specifier":",.1f"}},
        {"name": "Días cobertura",    "id": "Dias_Cobertura"},
        {"name": "Fecha quiebre",     "id": "Fecha_Quiebre"},
    ]

    # Detectar columnas de días (formato DD/MM)
    import re
    cols_dias = [c for c in df.columns if re.match(r"\d{2}/\d{2}", str(c))]
    cols_tabla = cols_fijas + [
        {"name": d, "id": d, "type":"numeric","format":{"specifier":",.0f"}}
        for d in cols_dias
    ]

    df["Semaforo_icon"] = df["Semaforo"].map({
        "rojo": "🔴", "amarillo": "🟡", "verde": "✅"})

    COLOR_SEM = {"rojo": "#ffc7ce", "amarillo": "#ffeb9c", "verde": "#c6efce"}
    # Umbrales: rojo < 7 días, amarillo < 15 días, verde >= 30 días
    style_cond = [
        {"if": {"filter_query": f'{{Semaforo}} = "{s}"'},
         "backgroundColor": c}
        for s, c in COLOR_SEM.items()
    ]

    tabla = dash_table.DataTable(
        id="tabla-proy",
        columns=cols_tabla,
        data=df.to_dict("records"),
        page_size=30,
        sort_action="native",
        filter_action="native",
        fixed_columns={"headers": True, "data": 4},
        style_table={"overflowX": "auto", "minWidth": "100%"},
        style_header={"backgroundColor": "#1f3864", "color": "white",
                      "fontWeight": "500", "fontSize": "11px",
                      "border": "0.5px solid #dee2e6", "textAlign": "center"},
        style_cell={"fontSize": "11px", "padding": "5px 7px",
                    "border": "0.5px solid #dee2e6", "textAlign": "center",
                    "minWidth": "70px"},
        style_data_conditional=style_cond,
        style_cell_conditional=[
            {"if": {"column_id": "Desc_Comp"}, "textAlign": "left", "minWidth": "200px"},
            {"if": {"column_id": "Fecha_Quiebre"},
             "fontWeight": "500", "color": "#c0392b"},
            {"if": {"column_id": "Dias_Cobertura"},
             "fontWeight": "500"},
        ],
    )
    return kpis, tabla, tipo_opts



@app.callback(
    Output("proy-feedback", "children"),
    Input("btn-export-proy", "n_clicks"),
    State("proy-fil-tipo", "value"),
    State("proy-fil-semaforo", "value"),
    prevent_initial_call=True,
)
def exportar_proyeccion(_, tipos, semaforos):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import re as re_mod

        df = get_mrp_proyectado(DB_PATH)
        if df.empty:
            return dbc.Alert("Sin datos para exportar.", color="warning", dismissable=True)

        if tipos:    df = df[df["Tipo_Material2"].isin(tipos)]
        if semaforos: df = df[df["Semaforo"].isin(semaforos)]

        wb = Workbook()
        ws = wb.active
        ws.title = "Proyección Quiebre Stock"

        s = Side(style="thin", color="BFBFBF")
        brd = Border(left=s, right=s, top=s, bottom=s)
        C_TITULO = "2F5496"
        C_AZUL   = "1F3864"
        COLOR_SEM = {"rojo":"FFC7CE","amarillo":"FFEB9C","verde":"C6EFCE"}

        # Título
        cols_dias = [c for c in df.columns if re_mod.match(r"\d{2}/\d{2}", str(c))]
        total_cols = 9 + len(cols_dias)
        ws.merge_cells(f"A1:{get_column_letter(total_cols)}1")
        ws["A1"] = f"Proyección de Quiebre de Stock – Plan completo"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
        ws["A1"].fill = PatternFill("solid", fgColor=C_TITULO)
        ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22

        # Encabezados fijos
        enc_fijos = ["Semáforo","Tipo Material","Código","Descripción",
                     "Stock inicial","Consumo total","Cons./día","Días cob.","Fecha quiebre"]
        id_fijos  = ["Semaforo","Tipo_Material2","Codigo_Comp","Desc_Comp",
                     "Stock_Inicial","Consumo_Total","Consumo_Diario_Prom","Dias_Cobertura","Fecha_Quiebre"]
        anchos_f  = [10,18,14,40,13,13,11,10,14]

        for c_idx, (enc, ancho) in enumerate(zip(enc_fijos, anchos_f), 1):
            cell = ws.cell(2, c_idx, enc)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor=C_AZUL)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = brd
            ws.column_dimensions[get_column_letter(c_idx)].width = ancho
        for d_idx, dia in enumerate(cols_dias, len(enc_fijos)+1):
            cell = ws.cell(2, d_idx, dia)
            cell.font = Font(bold=True, color="FFFFFF", size=8)
            cell.fill = PatternFill("solid", fgColor=C_AZUL)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = brd
            ws.column_dimensions[get_column_letter(d_idx)].width = 9
        ws.row_dimensions[2].height = 28

        # Filas
        for fi, (_, row) in enumerate(df.iterrows(), 3):
            sem = row["Semaforo"]
            bg = COLOR_SEM.get(sem, "FFFFFF")
            sem_icon = {"rojo":"🔴","amarillo":"🟡","verde":"✅"}.get(sem,"")
            vals_fijos = [sem_icon, row["Tipo_Material2"], row["Codigo_Comp"], row["Desc_Comp"],
                          int(row["Stock_Inicial"]), int(row["Consumo_Total"]),
                          round(float(row["Consumo_Diario_Prom"]),1),
                          row["Dias_Cobertura"], row["Fecha_Quiebre"]]
            for c_idx, val in enumerate(vals_fijos, 1):
                cell = ws.cell(fi, c_idx, val)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = brd
                cell.font = Font(size=9)
                cell.alignment = Alignment(
                    horizontal="left" if c_idx == 4 else "center", vertical="center")
                if c_idx in (5,6):
                    cell.number_format = "#,##0"
            for d_idx, dia in enumerate(cols_dias, len(enc_fijos)+1):
                val = row.get(dia, 0)
                cell = ws.cell(fi, d_idx, int(val) if val else 0)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = brd
                cell.font = Font(size=8)
                cell.alignment = Alignment(horizontal="center")
                cell.number_format = "#,##0"
            ws.row_dimensions[fi].height = 15

        ws.freeze_panes = f"{get_column_letter(len(enc_fijos)+1)}3"
        ws.auto_filter.ref = f"A2:{get_column_letter(total_cols)}{len(df)+2}"

        salida = "Proyeccion_Quiebre_Stock.xlsx"
        wb.save(salida)
        return dbc.Alert(f"✅ Exportado: {salida}", color="success",
                         dismissable=True, duration=5000)
    except Exception as e:
        return dbc.Alert(f"❌ {str(e)}", color="danger", dismissable=True)

@app.callback(
    Output("mrp-feedback", "children"),
    Input("btn-exportar-mrp", "n_clicks"),
    State("store-fecha", "data"),
    prevent_initial_call=True,
)
def exportar_mrp(_, fecha_str):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        fecha = date.fromisoformat(fecha_str)
        df = get_mrp_dia(DB_PATH, fecha)
        if df.empty:
            return dbc.Alert("Sin datos para exportar.", color="warning", dismissable=True)

        df["Alerta_label"] = df["Alerta"].apply(
            lambda x: "Falta stock" if x else "Cubierto")

        wb = Workbook()
        ws = wb.active
        ws.title = f"MRP {fecha.strftime('%d-%m-%Y')}"

        s = Side(style="thin", color="BFBFBF")
        brd = Border(left=s, right=s, top=s, bottom=s)

        # Título
        ws.merge_cells("A1:J1")
        ws["A1"] = f"MRP Diario – {fecha.strftime('%d/%m/%Y')}"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
        ws["A1"].fill = PatternFill("solid", fgColor="2F5496")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        cols = ["Proceso","Código comp.","Descripción","Req. bruto",
                "ALMA002","ALMA089","Stock prod.","Stock total","Req. neto","Cobertura"]
        ids  = ["Proceso","Codigo_Comp","Desc_Comp","Req_Bruto",
                "Stock_ALMA002","Stock_ALMA089","Stock_Prod","Stock_Total","Req_Neto","Alerta_label"]
        anchos = [14,16,42,13,12,12,12,12,12,14]

        for c_idx, (col, ancho) in enumerate(zip(cols, anchos), 1):
            cell = ws.cell(2, c_idx, col)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = brd
            ws.column_dimensions[get_column_letter(c_idx)].width = ancho
        ws.row_dimensions[2].height = 28

        for fi, (_, row) in enumerate(df.iterrows(), 3):
            es_alerta = row["Alerta"]
            bg = "FFC7CE" if es_alerta else "C6EFCE"
            for c_idx, col_id in enumerate(ids, 1):
                val = row[col_id]
                if col_id in ("Req_Bruto","Stock_ALMA002","Stock_ALMA089",
                              "Stock_Prod","Stock_Total","Req_Neto"):
                    val = int(val) if not pd.isna(val) else 0
                cell = ws.cell(fi, c_idx, val)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = brd
                cell.font = Font(size=9)
                cell.alignment = Alignment(
                    horizontal="left" if c_idx == 3 else "center")
                if c_idx in (4,5,6,7,8,9):
                    cell.number_format = "#,##0"

        ws.freeze_panes = "A3"
        ws.auto_filter.ref = f"A2:J{len(df)+2}"

        salida = f"MRP_{fecha.strftime('%Y%m%d')}.xlsx"
        wb.save(salida)
        return dbc.Alert(f"✅ Exportado: {salida}", color="success",
                         dismissable=True, duration=5000)
    except Exception as e:
        return dbc.Alert(f"❌ {str(e)}", color="danger", dismissable=True)


@app.callback(
    Output("export-feedback", "children"),
    Input("btn-exportar", "n_clicks"),
    prevent_initial_call=True,
)
def exportar_mps(_):
    try:
        result = subprocess.run(
            ["python3", "exportar_excel.py"],
            capture_output=True, text=True, timeout=60,
            cwd=os.path.dirname(os.path.abspath(__file__))
        )
        if result.returncode == 0:
            return dbc.Alert("✅ MPS_Reporte.xlsx exportado",
                             color="success", dismissable=True, duration=5000)
        return dbc.Alert(f"❌ {result.stderr[:300]}", color="danger", dismissable=True)
    except Exception as e:
        return dbc.Alert(f"❌ {str(e)}", color="danger", dismissable=True)


@app.callback(
    Output("modal", "is_open"),
    Output("m-feedback", "children"),
    Input("btn-abrir-modal", "n_clicks"),
    Input("btn-cancelar",    "n_clicks"),
    Input("btn-guardar",     "n_clicks"),
    State("modal", "is_open"),
    State("m-op",  "value"),
    State("m-qty", "value"),
    State("m-usuario", "value"),
    State("store-fecha", "data"),
    prevent_initial_call=True,
)
def manejar_modal(abrir, cancelar, guardar, is_open, op_id, qty, usuario, fecha_str):
    ctx = callback_context.triggered[0]["prop_id"]
    if "abrir"    in ctx: return True, ""
    if "cancelar" in ctx: return False, ""
    if "guardar"  in ctx:
        if not op_id or qty is None:
            return True, dbc.Alert("Completá la orden y la cantidad.",
                                   color="warning", className="py-1 px-2")
        fecha = date.fromisoformat(fecha_str)
        con = sqlite3.connect(DB_PATH)
        con.execute("""
            INSERT INTO avance_real (BELNR_ID, ItemCode, Descripcion, Fecha, Cantidad_Real)
            SELECT BELNR_ID, ItemCode, Descripcion, ?, ?
            FROM ordenes_plan WHERE BELNR_ID = ?
            ON CONFLICT(BELNR_ID, Fecha)
            DO UPDATE SET Cantidad_Real = excluded.Cantidad_Real
        """, (str(fecha), float(qty), int(op_id)))
        con.commit()
        con.close()
        return False, ""
    return is_open, ""



@app.callback(
    Output("cum-kpi-row",    "children"),
    Output("cum-panel-dias", "children"),
    Output("cum-panel-proc", "children"),
    Output("cum-tabla",      "children"),
    Output("cum-fil-proceso","options"),
    Input("btn-cum-refresh", "n_clicks"),
    Input("cum-fil-semana",  "value"),
    Input("cum-fil-proceso", "value"),
)
def actualizar_cumplimiento(_, semana_str, procesos_fil):
    if not semana_str:
        semana_str = str(date.today() - timedelta(days=date.today().weekday()))
    fecha = date.fromisoformat(semana_str)
    try:
        r = get_cumplimiento_detalle(DB_PATH, fecha)
    except Exception as e:
        err = dbc.Alert(f"Error: {str(e)}", color="danger")
        return [], err, err, err

    # Opciones de proceso
    proc_opts = [{"label": p["proceso"], "value": p["proceso"]}
                 for p in r["detalle_proceso"]]

    if not r["detalle_dia"]:
        msg = dbc.Alert("Sin datos para esta semana.", color="info")
        return [], msg, msg, msg, proc_opts

    # Filtrar órdenes por proceso si hay filtro activo
    ordenes_fil = r["detalle_ordenes"]
    if procesos_fil:
        ordenes_fil = [o for o in ordenes_fil if o["Proceso"] in procesos_fil]

    pct = r["pct_semana"]
    col = "success" if pct >= 90 else "warning" if pct >= 70 else "danger"

    # KPIs
    kpis = dbc.Row([
        dbc.Col(make_kpi(f"Semana {r['semana']}", f"{pct}%",
                         "cumplimiento", col), width=3),
        dbc.Col(make_kpi("Plan semanal", fmt_num(r["plan_total"]),
                         "unidades", "info"), width=3),
        dbc.Col(make_kpi("Real producido", fmt_num(r["real_total"]),
                         "unidades", "primary"), width=3),
        dbc.Col(make_kpi("Órdenes completas",
                         f"{r['ordenes_ok']} / {r['ordenes_total']}",
                         "completadas",
                         "success" if r["ordenes_ok"]==r["ordenes_total"] else "warning"),
                width=3),
    ], className="g-2")

    # Panel días
    def color_bar(p):
        if p >= 90: return "#1a7a4a"
        if p >= 70: return "#854F0B"
        return "#A32D2D"

    # Cumplimiento diario — recalcular por día si hay filtro de proceso
    barras_dia = []
    for d in r["detalle_dia"]:
        if procesos_fil:
            # Recalcular plan/real del día solo para las órdenes del proceso filtrado
            from db import get_ordenes_activas_en_dia as _get_dia
            try:
                fecha_d = date.fromisoformat(d["fecha"].split(" ")[-1].replace("/","") if "/" in d["fecha"] else str(fecha))
                # Parse fecha desde string "Lun 06/04"
                partes = d["fecha"].split(" ")
                if len(partes) == 2:
                    dm = partes[1].split("/")
                    fecha_d = date(fecha.year, int(dm[1]), int(dm[0]))
                df_d = _get_dia(DB_PATH, fecha_d)
                if not df_d.empty and procesos_fil:
                    df_d_fil = df_d[df_d["Proceso"].isin(procesos_fil)]
                    plan_d = df_d_fil["Plan_Dia"].sum()
                    real_d = df_d_fil["Real_Dia"].sum()
                    p = round(real_d / plan_d * 100, 1) if plan_d > 0 else 0
                else:
                    p = d["pct"]
            except Exception:
                p = d["pct"]
        else:
            p = d["pct"]
        bc = color_bar(p)
        barras_dia.append(
            html.Div(style={"display":"flex","alignItems":"center","gap":"8px",
                            "marginBottom":"10px"}, children=[
                html.Span(d["fecha"], style={"fontSize":"11px","width":"72px",
                           "textAlign":"right","color":"var(--color-text-secondary)","flexShrink":"0"}),
                html.Div(style={"flex":"1","height":"12px","background":"#e9ecef",
                                "borderRadius":"4px","overflow":"hidden"}, children=[
                    html.Div(style={"width":f"{min(p,100)}%","height":"100%",
                                    "background":bc,"borderRadius":"4px"})
                ]),
                html.Span(f"{p:.0f}%", style={"fontSize":"12px","fontWeight":"700",
                           "minWidth":"38px","textAlign":"right","color":bc}),
            ])
        )

    panel_dias = dbc.Card(dbc.CardBody([
        html.P("Cumplimiento diario", style={"fontSize":"12px","fontWeight":"500",
               "color":"var(--color-text-secondary)","marginBottom":"12px"}),
        *barras_dia,
    ]), style={"borderRadius":"10px","border":"0.5px solid var(--color-border-tertiary)",
               "background":"var(--color-background-secondary)"})

    # Panel proceso — dos barras: real acumulado y nota vs meta del día
    barras_proc = []
    for p_item in r["detalle_proceso"]:
        p_real     = p_item["pct"]             # % real vs plan total
        p_meta     = p_item.get("pct_vs_meta", 0)  # % real vs meta del día
        p_meta_dia = p_item.get("pct_meta_dia", 0)  # % del plan que debería estar hecho
        bc_real = color_bar(p_real)
        bc_meta = color_bar(p_meta)

        barras_proc.append(
            html.Div(style={"marginBottom":"12px"}, children=[
                # Nombre del proceso
                html.Div(style={"display":"flex","justifyContent":"space-between",
                                "marginBottom":"3px"}, children=[
                    html.Span(p_item["proceso"][:18],
                              style={"fontSize":"11px","fontWeight":"500",
                                     "color":"var(--color-text-primary)"}),
                    html.Span(f"Meta: {p_meta_dia:.0f}%",
                              style={"fontSize":"10px","color":"var(--color-text-tertiary)"}),
                ]),
                # Barra 1: avance real acumulado
                html.Div(style={"display":"flex","alignItems":"center","gap":"6px","marginBottom":"4px"}, children=[
                    html.Span("Real", style={"fontSize":"10px","width":"32px","color":"var(--color-text-secondary)","flexShrink":"0"}),
                    html.Div(style={"flex":"1","height":"10px","background":"#e9ecef",
                                    "borderRadius":"4px","overflow":"hidden"}, children=[
                        html.Div(style={"width":f"{min(p_real,100)}%","height":"100%",
                                        "background":bc_real,"borderRadius":"4px"})
                    ]),
                    html.Span(f"{p_real:.0f}%", style={"fontSize":"11px","fontWeight":"700",
                               "minWidth":"34px","textAlign":"right","color":bc_real}),
                ]),
                # Barra 2: nota vs meta del día
                html.Div(style={"display":"flex","alignItems":"center","gap":"6px"}, children=[
                    html.Span("Nota", style={"fontSize":"10px","width":"32px","color":"var(--color-text-secondary)","flexShrink":"0"}),
                    html.Div(style={"flex":"1","height":"10px","background":"#e9ecef",
                                    "borderRadius":"4px","overflow":"hidden"}, children=[
                        html.Div(style={"width":f"{min(p_meta,100)}%","height":"100%",
                                        "background":bc_meta,"borderRadius":"4px"})
                    ]),
                    html.Span(f"{p_meta:.0f}%", style={"fontSize":"11px","fontWeight":"700",
                               "minWidth":"34px","textAlign":"right","color":bc_meta}),
                ]),
            ])
        )

    panel_proc = dbc.Card(dbc.CardBody([
        html.Div(style={"display":"flex","justifyContent":"space-between","marginBottom":"12px"}, children=[
            html.P("Por proceso", style={"fontSize":"12px","fontWeight":"500",
                   "color":"var(--color-text-secondary)","margin":"0"}),
            html.P("Real = vs plan total  |  Nota = vs meta del día",
                   style={"fontSize":"10px","color":"var(--color-text-tertiary)","margin":"0"}),
        ]),
        *barras_proc,
    ]), style={"borderRadius":"10px","border":"0.5px solid var(--color-border-tertiary)",
               "background":"var(--color-background-secondary)"})

    # Tabla detalle órdenes
    ESTADO_COLOR = {
        "Completo":  "#c6efce", "Parcial": "#ffeb9c",
        "Atrasado":  "#ffc7ce", "Pendiente": "#f8f9fa",
    }
    ESTADO_TXT = {
        "Completo":  "#1a7a4a", "Parcial": "#854F0B",
        "Atrasado":  "#A32D2D", "Pendiente": "#888780",
    }

    tabla = dash_table.DataTable(
        id="tabla-cum",
        columns=[
            {"name":"O. Prod.",    "id":"BELNR_ID"},
            {"name":"Máquina",     "id":"Maquina"},
            {"name":"Proceso",     "id":"Proceso"},
            {"name":"Descripción", "id":"Descripcion"},
            {"name":"Planificado", "id":"Planificado",  "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Plan sem.",   "id":"plan_sem",     "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Real sem.",   "id":"real_sem",     "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"% Cum.",      "id":"pct",          "type":"numeric","format":{"specifier":".1f"}},
            {"name":"Estado",      "id":"estado"},
        ],
        data=ordenes_fil,
        page_size=40,
        sort_action="native",
        filter_action="native",
        style_table={"overflowX":"auto"},
        style_header={"backgroundColor":"#1f3864","color":"white",
                      "fontWeight":"500","fontSize":"11px",
                      "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_cell={"fontSize":"11px","padding":"6px 10px",
                    "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_data_conditional=[
            {"if":{"filter_query": '{estado} = "' + est + '"'},
             "backgroundColor": ESTADO_COLOR[est],
             "color": ESTADO_TXT[est]}
            for est in ESTADO_COLOR
        ],
        style_cell_conditional=[
            {"if":{"column_id":"Descripcion"},"textAlign":"left","minWidth":"180px"},
            {"if":{"column_id":"pct"},        "fontWeight":"700","fontSize":"12px"},
            {"if":{"column_id":"plan_sem"},   "fontWeight":"700","color":"#0C447C",
             "background":"#E6F1FB"},
            {"if":{"column_id":"real_sem"},   "fontWeight":"700","color":"#1a7a4a",
             "background":"#EAF3DE"},
        ],
    )

    return kpis, panel_dias, panel_proc, tabla, proc_opts


# ── Callback: Semiterminados proyección ───────────────────────────────────────

@app.callback(
    Output("semi-kpi-row",       "children"),
    Output("semi-tabla-container","children"),
    Output("semi-fil-tipo",      "options"),
    Input("btn-recalc-semi",     "n_clicks"),
    Input("semi-fil-tipo",       "value"),
    Input("semi-fil-semaforo",   "value"),
)
def actualizar_proyeccion_semi(_, tipos, semaforos):
    try:
        df = get_mrp_proyectado(DB_PATH)
    except Exception as e:
        return [], dbc.Alert(f"Error: {str(e)}", color="danger"), []

    if df.empty:
        return [], dbc.Alert("Sin datos.", color="info"), []

    # Solo semiterminados (231xxx)
    df = df[df["Codigo_Comp"].str.startswith("231")].copy()

    tipo_opts = [{"label": t, "value": t} for t in sorted(df["Tipo_Material2"].unique())]
    if tipos:    df = df[df["Tipo_Material2"].isin(tipos)]
    if semaforos: df = df[df["Semaforo"].isin(semaforos)]

    total      = len(df)
    n_rojo     = (df["Semaforo"] == "rojo").sum()
    n_amarillo = (df["Semaforo"] == "amarillo").sum()
    n_verde    = (df["Semaforo"] == "verde").sum()

    kpis = dbc.Row([
        dbc.Col(make_kpi("Semiterminados", str(total), "en el plan"), width=3),
        dbc.Col(make_kpi("🔴 Quiebre < 7 días",  str(int(n_rojo)),
                         "producir urgente", "danger" if n_rojo else "secondary"), width=3),
        dbc.Col(make_kpi("🟡 Quiebre < 15 días", str(int(n_amarillo)),
                         "planificar", "warning" if n_amarillo else "secondary"), width=3),
        dbc.Col(make_kpi("✅ Sin riesgo", str(int(n_verde)), "cobertura ok", "success"), width=3),
    ], className="g-2")

    import re
    cols_dias = [c for c in df.columns if re.match(r"\d{2}/\d{2}", str(c))]
    df["Semaforo_icon"] = df["Semaforo"].map({"rojo":"🔴","amarillo":"🟡","verde":"✅"})
    COLOR_SEM = {"rojo":"#ffc7ce","amarillo":"#ffeb9c","verde":"#c6efce"}

    cols_tabla = [
        {"name":"Semáforo",      "id":"Semaforo_icon"},
        {"name":"Tipo",          "id":"Tipo_Material2"},
        {"name":"Código",        "id":"Codigo_Comp"},
        {"name":"Descripción",   "id":"Desc_Comp"},
        {"name":"Stock inicial", "id":"Stock_Inicial",       "type":"numeric","format":{"specifier":",.0f"}},
        {"name":"Consumo total", "id":"Consumo_Total",       "type":"numeric","format":{"specifier":",.0f"}},
        {"name":"Diferencia",    "id":"Diferencia",          "type":"numeric","format":{"specifier":",.0f"}},
        {"name":"Cons./día",     "id":"Consumo_Diario_Prom", "type":"numeric","format":{"specifier":",.1f"}},
        {"name":"Días cob.",     "id":"Dias_Cobertura"},
        {"name":"Fecha quiebre", "id":"Fecha_Quiebre"},
    ] + [{"name": d, "id": d, "type":"numeric","format":{"specifier":",.0f"}} for d in cols_dias]

    tabla = dash_table.DataTable(
        id="tabla-semi",
        columns=cols_tabla,
        data=df.to_dict("records"),
        page_size=30, sort_action="native", filter_action="native",
        fixed_columns={"headers": True, "data": 4},
        style_table={"overflowX":"auto","minWidth":"100%"},
        style_header={"backgroundColor":"#1f3864","color":"white",
                      "fontWeight":"500","fontSize":"11px","border":"0.5px solid #dee2e6"},
        style_cell={"fontSize":"11px","padding":"5px 7px",
                    "border":"0.5px solid #dee2e6","textAlign":"center","minWidth":"70px"},
        style_data_conditional=[
            {"if":{"filter_query":f'{{Semaforo}} = "{s}"'},"backgroundColor":c}
            for s,c in COLOR_SEM.items()
        ],
        style_cell_conditional=[
            {"if":{"column_id":"Desc_Comp"},"textAlign":"left","minWidth":"200px"},
            {"if":{"column_id":"Fecha_Quiebre"},"fontWeight":"500","color":"#c0392b"},
        ],
    )
    return kpis, tabla, tipo_opts


@app.callback(
    Output("semi-feedback", "children"),
    Input("btn-export-semi", "n_clicks"),
    prevent_initial_call=True,
)
def exportar_semi(_):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        import re as re_mod

        df = get_mrp_proyectado(DB_PATH)
        df = df[df["Codigo_Comp"].str.startswith("231")].copy()
        if df.empty:
            return dbc.Alert("Sin datos.", color="warning", dismissable=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Proyección Semiterminados"
        s = Side(style="thin", color="BFBFBF")
        brd = Border(left=s, right=s, top=s, bottom=s)
        COLOR_SEM = {"rojo":"FFC7CE","amarillo":"FFEB9C","verde":"C6EFCE"}

        cols_dias = [c for c in df.columns if re_mod.match(r"\d{2}/\d{2}", str(c))]
        enc = ["Semáforo","Tipo","Código","Descripción","Stock ini.","Consumo total",
               "Cons./día","Días cob.","Fecha quiebre"] + cols_dias
        ids = ["Semaforo","Tipo_Material2","Codigo_Comp","Desc_Comp","Stock_Inicial",
               "Consumo_Total","Consumo_Diario_Prom","Dias_Cobertura","Fecha_Quiebre"] + cols_dias

        ws.merge_cells(f"A1:{get_column_letter(len(enc))}1")
        ws["A1"] = "Proyección de Semiterminados – Quiebre de Stock"
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=12)
        ws["A1"].fill = PatternFill("solid", fgColor="2F5496")
        ws["A1"].alignment = Alignment(horizontal="center")
        ws.row_dimensions[1].height = 22

        for ci, e in enumerate(enc, 1):
            cell = ws.cell(2, ci, e)
            cell.font = Font(bold=True, color="FFFFFF", size=9)
            cell.fill = PatternFill("solid", fgColor="1F3864")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = brd
            ws.column_dimensions[get_column_letter(ci)].width = 10 if e in cols_dias else 14
        ws.row_dimensions[2].height = 28

        for fi, (_, row) in enumerate(df.iterrows(), 3):
            sem = row["Semaforo"]
            bg = COLOR_SEM.get(sem, "FFFFFF")
            icon = {"rojo":"🔴","amarillo":"🟡","verde":"✅"}.get(sem,"")
            vals = [icon, row["Tipo_Material2"], row["Codigo_Comp"], row["Desc_Comp"],
                    int(row["Stock_Inicial"]), int(row["Consumo_Total"]),
                    round(float(row["Consumo_Diario_Prom"]),1),
                    row["Dias_Cobertura"], row["Fecha_Quiebre"]]
            vals += [int(row.get(d,0)) for d in cols_dias]
            for ci, val in enumerate(vals, 1):
                cell = ws.cell(fi, ci, val)
                cell.fill = PatternFill("solid", fgColor=bg)
                cell.border = brd
                cell.font = Font(size=9)
                cell.alignment = Alignment(
                    horizontal="left" if ci == 4 else "center")
            ws.row_dimensions[fi].height = 15

        ws.freeze_panes = "J3"
        ws.auto_filter.ref = f"A2:{get_column_letter(len(enc))}{len(df)+2}"
        wb.save("Proyeccion_Semiterminados.xlsx")
        return dbc.Alert("✅ Exportado: Proyeccion_Semiterminados.xlsx",
                         color="success", dismissable=True, duration=5000)
    except Exception as e:
        return dbc.Alert(f"❌ {str(e)}", color="danger", dismissable=True)


# ── Callback: Cumplimiento con filtros de semana y proceso ────────────────────

@app.callback(
    Output("cum-fil-semana", "options"),
    Output("cum-fil-semana", "value"),
    Input("tabs", "active_tab"),
)
def cargar_semanas(tab):
    semanas = get_semanas_plan(DB_PATH)
    # Seleccionar semana actual por defecto
    hoy = date.today()
    lunes_hoy = str(hoy - timedelta(days=hoy.weekday()))
    val_default = lunes_hoy if any(s["value"] == lunes_hoy for s in semanas) else (semanas[0]["value"] if semanas else None)
    return semanas, val_default


# ── Callback: Gantt dashboard ─────────────────────────────────────────────────

@app.callback(
    Output("gantt-fil-mes", "options"),
    Output("gantt-fil-proceso", "options"),
    Output("gantt-fil-maquina", "options"),
    Output("gantt-fil-linea", "options"),
    Input("tabs", "active_tab"),
    Input("auto-refresh", "n_intervals"),
)
def cargar_opciones_gantt(tab, _):
    try:
        import sqlite3 as _sq
        con = _sq.connect(DB_PATH)
        df = pd.read_sql("SELECT DISTINCT Proceso, Maquina, Linea, Fecha_Inicio, Fecha_Fin FROM ordenes_plan", con)
        con.close()

        proc_opts = [{"label": p, "value": p} for p in sorted(df["Proceso"].unique())]
        maq_opts  = [{"label": m, "value": m} for m in sorted(df["Maquina"].unique())]

        # Meses disponibles en el plan
        meses = set()
        for _, row in df.iterrows():
            for col in ["Fecha_Inicio", "Fecha_Fin"]:
                d = date.fromisoformat(str(row[col])[:10])
                meses.add((d.year, d.month))
        import calendar
        mes_opts = [
            {"label": f"{calendar.month_name[m]} {y}", "value": f"{y}-{m:02d}"}
            for y, m in sorted(meses)
        ]
        linea_opts = [{"label": l, "value": l} for l in sorted(df["Linea"].dropna().unique()) if l]
        return mes_opts, proc_opts, maq_opts, linea_opts
    except Exception:
        return [], [], [], []


@app.callback(
    Output("gantt-container", "children"),
    Input("btn-gantt-load", "n_clicks"),
    State("gantt-fil-proceso", "value"),
    State("gantt-fil-mes", "value"),
    State("gantt-fil-maquina", "value"),
    State("gantt-fil-linea", "value"),
    prevent_initial_call=True,
)
def actualizar_gantt(_, proceso, mes_str, maquinas, lineas):
    import sqlite3 as _sq
    from db import es_habil, get_todos_ingresos

    if not mes_str:
        return dbc.Alert("Seleccioná un mes.", color="warning")

    year, month = int(mes_str.split("-")[0]), int(mes_str.split("-")[1])
    import calendar
    primer_dia = date(year, month, 1)
    ultimo_dia = date(year, month, calendar.monthrange(year, month)[1])
    dias = [primer_dia + timedelta(days=i)
            for i in range((ultimo_dia - primer_dia).days + 1)
            if es_habil(primer_dia + timedelta(days=i))]

    if not dias:
        return dbc.Alert("Sin días hábiles en ese mes.", color="info")

    con = _sq.connect(DB_PATH)
    df_plan = pd.read_sql("SELECT * FROM ordenes_plan", con)
    con.close()
    df_ingr = get_todos_ingresos(DB_PATH)

    # Cargar mantenimientos
    try:
        df_mant = pd.read_excel(EXCEL_PLAN, sheet_name="Mantenimientos")
        df_mant.columns = df_mant.columns.str.strip()
        df_mant["Fecha_Inicio"] = pd.to_datetime(df_mant["Fecha_Inicio"]).dt.date
        df_mant["Fecha_Fin"]    = pd.to_datetime(df_mant["Fecha_Fin"]).dt.date
    except Exception:
        df_mant = pd.DataFrame(columns=["Maquina","Fecha_Inicio","Fecha_Fin","Descripcion"])

    # Índice mantenimientos: {(maquina, fecha): descripcion}
    mant_idx = {}
    for _, mr in df_mant.iterrows():
        d = mr["Fecha_Inicio"]
        while d <= mr["Fecha_Fin"]:
            mant_idx[(str(mr["Maquina"]).strip(), str(d))] = str(mr["Descripcion"])
            d += timedelta(days=1)

    # Filtros
    if proceso:
        df_plan = df_plan[df_plan["Proceso"] == proceso]
    if maquinas:
        df_plan = df_plan[df_plan["Maquina"].isin(maquinas)]
    if lineas:
        df_plan = df_plan[df_plan["Linea"].isin(lineas)]

    # Solo OPs activas en el mes
    df_plan = df_plan[
        (df_plan["Fecha_Inicio"] <= str(ultimo_dia)) &
        (df_plan["Fecha_Fin"]    >= str(primer_dia))
    ].reset_index(drop=True)

    if df_plan.empty:
        return dbc.Alert("No hay órdenes activas en este mes/proceso.", color="info")

    # Índice ingresos
    ingr_dia = {}
    if not df_ingr.empty and "Cantidad_Real" in df_ingr.columns:
        for _, r in df_ingr.iterrows():
            ingr_dia[(int(r["BELNR_ID"]), str(r["Fecha"])[:10])] = float(r["Cantidad_Real"])

    # Construir tabla
    COLS_FIJAS = ["OP", "Máquina", "Descripción", "Planificado"]
    cols_header = COLS_FIJAS + [d.strftime("%d/%m") for d in dias]

    rows_data = []
    for _, row in df_plan.iterrows():
        belnr    = int(row["BELNR_ID"])
        f_inicio = date.fromisoformat(str(row["Fecha_Inicio"])[:10])
        f_fin    = date.fromisoformat(str(row["Fecha_Fin"])[:10])
        cap_dia  = float(row["Cap_Diaria"])
        planif   = float(row["Planificado"])

        fila = {
            "OP": belnr,
            "Máquina": row["Maquina"],
            "Descripción": row["Descripcion"][:35],
            "Planificado": f"{int(planif):,}",
        }

        acum_plan = 0.0
        acum_real = 0.0

        for dia in dias:
            key = dia.strftime("%d/%m")
            real_dia = ingr_dia.get((belnr, dia.strftime("%Y-%m-%d")), None)

            # Mantenimiento — celda verde con descripción
            mant_desc = mant_idx.get((str(row["Maquina"]).strip(), str(dia)))
            if mant_desc:
                fila[key] = f"🔧 {mant_desc}"
                continue

            if dia < f_inicio:
                fila[key] = ""
                continue

            if dia > f_fin:
                # Fuera del rango teórico — mostrar solo si hay real registrado
                if real_dia is not None:
                    acum_real += real_dia
                    fila[key] = "R:" + f"{int(real_dia):,}" + "\nA:" + f"{int(acum_real):,}"
                else:
                    fila[key] = ""
                continue

            restante = planif - acum_plan
            plan_dia = min(cap_dia, max(restante, 0))
            acum_plan += plan_dia
            if real_dia is not None:
                acum_real += real_dia

            deberia = acum_plan
            if real_dia is None:
                fila[key] = "P:" + f"{int(plan_dia):,}" + "\nD:" + f"{int(deberia):,}"
            else:
                pct = real_dia / plan_dia * 100 if plan_dia > 0 else 0
                fila[key] = (
                    "P:" + f"{int(plan_dia):,}" + "\n"
                    + "R:" + f"{int(real_dia):,}" + "\n"
                    + "A:" + f"{int(acum_real):,}" + "\n"
                    + "D:" + f"{int(deberia):,}" + "\n"
                    + f"{pct:.0f}%"
                )

        rows_data.append(fila)

    # Colores semáforo por celda + mantenimientos
    COLOR_GANTT = []
    for i, row_g in enumerate(rows_data):
        maquina_g = row_g.get("Máquina","")
        for dia in dias:
            key = dia.strftime("%d/%m")
            val = row_g.get(key, "")
            # Mantenimiento: celda verde
            if (maquina_g, str(dia)) in mant_idx:
                COLOR_GANTT.append({
                    "if": {"row_index": i, "column_id": key},
                    "backgroundColor": "#90EE90", "color": "#1a5c1a",
                })
                continue
            if not val or val == "":
                continue
            if "R:" in val:
                lines = val.split("\n")
                try:
                    pct_line = [l for l in lines if "%" in l]
                    pct = float(pct_line[0].replace("%","").split("/")[0]) if pct_line else 0
                    color = "#c6efce" if pct >= 85 else "#ffeb9c" if pct >= 65 else "#ffc7ce"
                except Exception:
                    color = "#f8f9fa"
            else:
                color = "#f0f0f0"
            COLOR_GANTT.append({
                "if": {"row_index": i, "column_id": key},
                "backgroundColor": color,
            })

    tabla = dash_table.DataTable(
        id="tabla-gantt",
        columns=[{"name": c, "id": c} for c in cols_header],
        data=rows_data,
        style_table={"overflowX": "auto", "minWidth": "100%", "width": "100%"},
        style_header={"backgroundColor":"#1f3864","color":"white",
                      "fontWeight":"500","fontSize":"11px",
                      "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_cell={"fontSize":"11px","padding":"6px 8px",
                    "border":"0.5px solid #dee2e6","textAlign":"center",
                    "whiteSpace":"pre-line","minWidth":"110px","maxWidth":"140px"},
        style_cell_conditional=[
            {"if":{"column_id":"Descripción"},"textAlign":"left",
             "minWidth":"220px","maxWidth":"280px"},
            {"if":{"column_id":"OP"},"minWidth":"70px","maxWidth":"80px"},
            {"if":{"column_id":"Máquina"},"minWidth":"80px","maxWidth":"90px"},
            {"if":{"column_id":"Planificado"},"minWidth":"90px","maxWidth":"100px"},
        ],
        style_data_conditional=COLOR_GANTT,
        page_size=30,
        fixed_columns={"headers":True,"data":4},
    )

    titulo = f"{'  |  '.join([proceso or 'Todos', mes_str])} — {len(df_plan)} órdenes"
    return html.Div([
        html.P(titulo, style={"fontSize":"12px","color":"var(--color-text-secondary)",
                               "marginBottom":"8px"}),
        tabla,
        html.Div(style={"marginTop":"8px","fontSize":"11px","color":"var(--color-text-secondary)"},
                 children="P: Plan día  |  R: Real día  |  A: Acum. real  |  D: Debería ir  |  %: cumplimiento")
    ])


# ── Callback: Resumen OPs ─────────────────────────────────────────────────────

@app.callback(
    Output("res-kpi-row",    "children"),
    Output("res-tabla",      "children"),
    Output("res-fil-proceso","options"),
    Output("res-fil-linea",  "options"),
    Input("btn-res-refresh", "n_clicks"),
    Input("auto-refresh",    "n_intervals"),
    Input("res-fil-proceso", "value"),
    Input("res-fil-linea",   "value"),
    Input("res-fil-estado",  "value"),
    Input("res-fil-op",      "value"),
)
def actualizar_resumen(_, __, procesos, lineas, estados, buscar_op):
    df = get_resumen_ops(DB_PATH)
    if df.empty:
        return [], dbc.Alert("Sin datos.", color="info"), [], []

    proc_opts  = [{"label":p,"value":p} for p in sorted(df["Proceso"].unique())]
    linea_opts = [{"label":l,"value":l} for l in sorted(df["Linea"].dropna().unique()) if l]

    if procesos:  df = df[df["Proceso"].isin(procesos)]
    if lineas:    df = df[df["Linea"].isin(lineas)]
    if estados:   df = df[df["Estado"].isin(estados)]
    if buscar_op: df = df[df["BELNR_ID"].astype(str).str.contains(buscar_op.strip())]

    total        = len(df)
    completadas  = (df["Estado"] == "Completado").sum()
    en_curso     = (df["Estado"] == "En curso").sum()
    pendientes   = (df["Estado"] == "Pendiente").sum()

    kpis = dbc.Row([
        dbc.Col(make_kpi("Total OPs", str(total), "en el plan"), width=3),
        dbc.Col(make_kpi("Completadas", str(int(completadas)), "finalizadas", "success"), width=3),
        dbc.Col(make_kpi("En curso", str(int(en_curso)), "activas", "info"), width=3),
        dbc.Col(make_kpi("Pendientes", str(int(pendientes)), "sin iniciar", "secondary"), width=3),
    ], className="g-2")

    ESTADO_COLOR = {
        "Completado": "#c6efce", "En curso": "#E6F1FB",
        "Parcial":    "#ffeb9c", "Pendiente": "#f8f9fa",
    }
    ESTADO_TXT = {
        "Completado": "#1a7a4a", "En curso": "#0C447C",
        "Parcial":    "#854F0B", "Pendiente": "#888780",
    }

    tabla = dash_table.DataTable(
        id="tabla-resumen",
        columns=[
            {"name":"O. Prod.",    "id":"BELNR_ID"},
            {"name":"Proceso",     "id":"Proceso"},
            {"name":"Línea",       "id":"Linea"},
            {"name":"Máquina",     "id":"Maquina"},
            {"name":"Sec",         "id":"Sec"},
            {"name":"Código",      "id":"ItemCode"},
            {"name":"Descripción", "id":"Descripcion"},
            {"name":"Inicio",      "id":"Inicio"},
            {"name":"Fin",         "id":"Fin"},
            {"name":"Planificado", "id":"Planificado",  "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Acumulado",   "id":"Acumulado",    "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Pendiente",   "id":"Pendiente",    "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"% Acum.",     "id":"Pct_Acum",     "type":"numeric","format":{"specifier":".1f"}},
            {"name":"Estado",      "id":"Estado"},
        ],
        data=df.to_dict("records"),
        page_size=50,
        sort_action="native",
        filter_action="native",
        style_table={"overflowX":"auto"},
        style_header={"backgroundColor":"#1f3864","color":"white",
                      "fontWeight":"500","fontSize":"11px",
                      "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_cell={"fontSize":"11px","padding":"6px 8px",
                    "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_data_conditional=[
            {"if":{"filter_query": '{Estado} = "' + est + '"'},
             "backgroundColor": ESTADO_COLOR[est], "color": ESTADO_TXT[est]}
            for est in ESTADO_COLOR
        ],
        style_cell_conditional=[
            {"if":{"column_id":"Descripcion"},"textAlign":"left","minWidth":"200px"},
            {"if":{"column_id":"Pendiente"},  "fontWeight":"500","color":"#c0392b"},
            {"if":{"column_id":"Acumulado"},  "fontWeight":"500","color":"#1a7a4a"},
            {"if":{"column_id":"Pct_Acum"},   "fontWeight":"700"},
        ],
    )
    return kpis, tabla, proc_opts, linea_opts


# ── Callback: Avance campaña ──────────────────────────────────────────────────

@app.callback(
    Output("camp-kpi-row",     "children"),
    Output("camp-tabla",       "children"),
    Output("camp-fil-linea",   "options"),
    Output("camp-fil-proceso", "options"),
    Output("camp-fil-maquina", "options"),
    Input("btn-camp-refresh",  "n_clicks"),
    Input("store-fecha",       "data"),
    Input("auto-refresh",      "n_intervals"),
    Input("camp-fil-linea",    "value"),
    Input("camp-fil-estado",   "value"),
    Input("camp-fil-proceso",  "value"),
    Input("camp-fil-maquina",  "value"),
)
def actualizar_campana(_, fecha_str, __, lineas, estados, procesos, maquinas):
    fecha = date.fromisoformat(fecha_str)
    df = get_avance_campana(DB_PATH, fecha)

    linea_opts = []
    proc_opts  = []
    maq_opts   = []
    if not df.empty:
        linea_opts = [{"label":l,"value":l} for l in sorted(df["Linea"].dropna().unique()) if l and str(l).strip()]
        if "Proceso" in df.columns:
            proc_opts = [{"label":p,"value":p} for p in sorted(df["Proceso"].dropna().unique()) if p and str(p).strip()]
        maq_opts = [{"label":m,"value":m} for m in sorted(df["Maquina"].dropna().unique()) if m]

    if df.empty:
        return [], dbc.Alert("Sin maquinas activas para esta fecha.", color="info"), linea_opts, proc_opts, maq_opts

    if procesos: df = df[df["Proceso"].isin(procesos)]
    if maquinas: df = df[df["Maquina"].isin(maquinas)]
    if lineas:   df = df[df["Linea"].isin(lineas)]
    if estados:  df = df[df["Estado"].isin(estados)]

    # KPIs
    total    = len(df)
    criticos = (df["Estado"] == "Critico").sum()
    atraso   = (df["Estado"] == "Retraso").sum()
    al_dia   = (df["Estado"] == "Al dia").sum()

    kpis = dbc.Row([
        dbc.Col(make_kpi("Máquinas activas", str(total), "hoy"), width=3),
        dbc.Col(make_kpi("🔴 Crítico",  str(int(criticos)),
                         "atraso > 3 días", "danger"  if criticos else "secondary"), width=3),
        dbc.Col(make_kpi("🟠 Retraso",  str(int(atraso)),
                         "1-3 días atrás",  "warning" if atraso   else "secondary"), width=3),
        dbc.Col(make_kpi("✅ Al día",   str(int(al_dia)),
                         "en ritmo",        "success" if al_dia   else "secondary"), width=3),
    ], className="g-2")

    # Colores por estado
    ESTADO_COLOR = {
        "Critico":      "#ffc7ce",
        "Retraso":      "#FFD580",
        "Leve retraso": "#ffeb9c",
        "Al dia":       "#c6efce",
    }
    ESTADO_TXT = {
        "Critico":      "#A32D2D",
        "Retraso":      "#7B4200",
        "Leve retraso": "#854F0B",
        "Al dia":       "#1a7a4a",
    }

    tabla = dash_table.DataTable(
        id="tabla-campana",
        columns=[
            {"name":"Proceso",      "id":"Proceso"},
            {"name":"Máquina",      "id":"Maquina"},
            {"name":"Planificado",  "id":"Planificado",  "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Debería ir",   "id":"Deberia_Ir",   "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Acumulado",    "id":"Acumulado",    "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Diferencia",   "id":"Diferencia",   "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"% Acum.",      "id":"Pct_Acum",     "type":"numeric","format":{"specifier":".1f"}},
            {"name":"% Debería",    "id":"Pct_Deberia",  "type":"numeric","format":{"specifier":".1f"}},
            {"name":"Días atraso",  "id":"Dias_Atraso"},
            {"name":"Cap. diaria",  "id":"Cap_Diaria",   "type":"numeric","format":{"specifier":",.0f"}},
            {"name":"Días trans.",  "id":"Dias_Transcurridos"},
            {"name":"Estado",       "id":"Estado"},
        ],
        data=df.to_dict("records"),
        sort_action="native",
        style_table={"overflowX":"auto"},
        style_header={"backgroundColor":"#1f3864","color":"white",
                      "fontWeight":"500","fontSize":"11px",
                      "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_cell={"fontSize":"11px","padding":"6px 8px",
                    "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_data_conditional=[
            {"if":{"filter_query":'{Estado} = "' + est + '"'},
             "backgroundColor": ESTADO_COLOR[est], "color": ESTADO_TXT[est]}
            for est in ESTADO_COLOR
        ],
        style_cell_conditional=[
            {"if":{"column_id":"Diferencia"},
             "fontWeight":"700"},
            {"if":{"column_id":"Acumulado"},
             "fontWeight":"500","color":"#1a7a4a"},
            {"if":{"column_id":"Deberia_Ir"},
             "fontWeight":"500","color":"#0C447C"},
            {"if":{"column_id":"Dias_Atraso"},
             "fontWeight":"700"},
            {"if":{"column_id":"Resumen"},
             "fontWeight":"700","textAlign":"left","minWidth":"180px"},
        ],
        page_size=30,
    )
    return kpis, tabla, linea_opts, proc_opts, maq_opts


# ── Callback: Descargar cumplimiento semanal ──────────────────────────────────

@app.callback(
    Output("cum-download", "data"),
    Input("btn-cum-descargar", "n_clicks"),
    State("cum-fil-semana",   "value"),
    State("cum-fil-proceso",  "value"),
    prevent_initial_call=True,
)
def descargar_cumplimiento_semanal(n, semana_vals, procesos):
    if not semana_vals:
        return None

    from datetime import date, timedelta
    import sqlite3 as _sq
    import io
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from db import es_habil as es_habil

    # Normalizar a lista
    if isinstance(semana_vals, str):
        semana_vals = [semana_vals]

    con = _sq.connect(DB_PATH)
    df_plan = pd.read_sql("SELECT * FROM ordenes_plan", con)
    con.close()

    all_rows = []
    for semana_str in sorted(semana_vals):
        lunes_sem   = date.fromisoformat(semana_str)
        viernes_sem = lunes_sem + timedelta(days=4)
        semana_num  = lunes_sem.isocalendar()[1]

        con = _sq.connect(DB_PATH)
        df_av = pd.read_sql(
            "SELECT BELNR_ID, SUM(Cantidad_Real) as real_sem FROM avance_real "
            "WHERE Fecha >= ? AND Fecha <= ? GROUP BY BELNR_ID",
            con, params=[str(lunes_sem), str(viernes_sem)])
        df_acum = pd.read_sql(
            "SELECT BELNR_ID, SUM(Cantidad_Real) as acum FROM avance_real "
            "GROUP BY BELNR_ID", con)
        con.close()

        av_map   = dict(zip(df_av["BELNR_ID"].astype(int),   df_av["real_sem"].fillna(0)))
        acum_map = dict(zip(df_acum["BELNR_ID"].astype(int), df_acum["acum"].fillna(0)))

        df_sem = df_plan[
            (df_plan["Fecha_Inicio"] <= str(viernes_sem)) &
            (df_plan["Fecha_Fin"]    >= str(lunes_sem))
        ].reset_index(drop=True)

        if procesos:
            df_sem = df_sem[df_sem["Proceso"].isin(procesos)]

        for _, row in df_sem.iterrows():
            f_ini_op = date.fromisoformat(str(row["Fecha_Inicio"])[:10])
            f_fin_op = date.fromisoformat(str(row["Fecha_Fin"])[:10])
            cap      = float(row["Cap_Diaria"])
            planif   = float(row["Planificado"])

            inicio_sem = max(f_ini_op, lunes_sem)
            fin_sem    = min(f_fin_op, viernes_sem)

            acum_prev = 0.0
            d = f_ini_op
            while d < inicio_sem:
                if es_habil(d):
                    acum_prev += min(cap, max(planif - acum_prev, 0))
                d += timedelta(days=1)

            plan_sem = 0.0
            d = inicio_sem
            while d <= fin_sem:
                if es_habil(d):
                    dia_plan  = min(cap, max(planif - acum_prev, 0))
                    plan_sem  += dia_plan
                    acum_prev += dia_plan
                d += timedelta(days=1)

            real_sem   = float(av_map.get(int(row["BELNR_ID"]), 0))
            acum_real  = float(acum_map.get(int(row["BELNR_ID"]), 0))
            avance_sap = float(row.get("Avance_SAP", 0)) if "Avance_SAP" in row.index else 0
            pct        = round(real_sem / plan_sem * 100, 1) if plan_sem > 0 else 0

            all_rows.append({
                "Semana":       semana_num,
                "Lunes":        lunes_sem.strftime("%d/%m/%Y"),
                "Viernes":      viernes_sem.strftime("%d/%m/%Y"),
                "O. Prod.":     int(row["BELNR_ID"]),
                "Mes":          int(row.get("Mes", 0)),
                "Máquina":      row["Maquina"],
                "Proceso":      row["Proceso"],
                "Sec":          int(row["Sec"]),
                "Código":       row["ItemCode"],
                "Descripción":  row["Descripcion"],
                "Planificado":  round(planif),
                "Avance SAP":   round(avance_sap),
                "Cap. diaria":  round(cap),
                "Plan semana":  round(plan_sem),
                "Real semana":  round(real_sem),
                "Acum. real":   round(acum_real),
                "% Cumpl.":     pct,
            })

    if not all_rows:
        return None

    df_out = pd.DataFrame(all_rows)
    semanas_txt = "_".join([str(date.fromisoformat(s).isocalendar()[1]) for s in sorted(semana_vals)])

    s = Side(style="thin", color="BFBFBF")
    brd = Border(left=s, right=s, top=s, bottom=s)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Semanas {semanas_txt}"[:31]

    cols = [
        ("Semana",7),("Lunes",10),("Viernes",10),("O. Prod.",9),("Mes",5),
        ("Máquina",9),("Proceso",14),("Sec",5),("Código",14),
        ("Descripción",40),("Planificado",13),("Avance SAP",13),("Cap. diaria",13),
        ("Plan semana",13),("Real semana",13),("Acum. real",13),("% Cumpl.",10),
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    sem_rango = f"Semanas {semanas_txt}"
    ws["A1"] = f"Cumplimiento {sem_rango}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill      = PatternFill("solid", fgColor="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c_idx, (col_name, ancho) in enumerate(cols, 1):
        cell = ws.cell(2, c_idx, col_name)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = brd
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws.row_dimensions[2].height = 28

    for fi, (_, row) in enumerate(df_out.iterrows(), 3):
        vals = [row[c[0]] for c in cols]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(fi, c_idx, val)
            cell.border    = brd
            cell.font      = Font(size=9)
            cell.alignment = Alignment(
                horizontal="left" if c_idx == 10 else "center",
                vertical="center")
            if c_idx in (11, 12, 13, 14, 15, 16):
                cell.number_format = "#,##0"
            if c_idx == 17:
                cell.number_format = "0.0"
                pct_val = float(val) if val else 0
                if   pct_val >= 90: cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif pct_val >= 70: cell.fill = PatternFill("solid", fgColor="FFEB9C")
                else:               cell.fill = PatternFill("solid", fgColor="FFC7CE")
        ws.row_dimensions[fi].height = 15

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(df_out)+2}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nombre = f"Cumplimiento_Semanas{semanas_txt}.xlsx"
    return dcc.send_bytes(buf.read(), nombre)



# ── Callback: Resumen Campaña ejecutivo ──────────────────────────────────────

@app.callback(
    Output("cumres-kpi-row",    "children"),
    Output("cumres-graficos",   "children"),
    Output("cumres-tabla",      "children"),
    Output("cumres-fil-semana", "options"),
    Output("cumres-fil-mes",    "options"),
    Output("cumres-fil-linea",  "options"),
    Output("cumres-fil-sub",    "options"),
    Input("btn-cumres-refresh", "n_clicks"),
    Input("cum-subtabs",        "active_tab"),
    Input("cumres-fil-semana",  "value"),
    Input("cumres-fil-mes",     "value"),
    Input("cumres-fil-linea",   "value"),
    Input("cumres-fil-sub",     "value"),
)
def actualizar_resumen_campana(_, tab, semanas, mes_sel, lineas, subs):
    import plotly.graph_objects as go

    try:
        df = pd.read_excel(EXCEL_PLAN, sheet_name="Resumen_Campaña")
        df.columns = df.columns.str.strip()
    except Exception as e:
        empty = [[], dbc.Alert(f"No se encontró Resumen_Campaña: {e}", color="warning"), [], [], [], [], []]
        return empty

    # Detectar columnas de meses dinámicamente
    cols_fijas = ["Proceso_Interno","Línea","Subcomponente","Plan Campaña"]
    plan_cols  = [c for c in df.columns if c not in cols_fijas and "Plan" in c]
    real_cols  = [c for c in df.columns if c not in cols_fijas and ("Real" in c or "real" in c.lower())]

    # Limpiar números
    for c in plan_cols + real_cols + (["Plan Campaña"] if "Plan Campaña" in df.columns else []):
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    # Opciones filtros — no hay Semana/Mes en esta estructura
    mes_map    = {4:"Abril",5:"Mayo",6:"Junio",7:"Julio",8:"Agosto",9:"Septiembre",10:"Octubre",11:"Noviembre",12:"Diciembre"}
    if "Semana" in df.columns:
        df["Semana"] = pd.to_numeric(df["Semana"], errors="coerce").fillna(0).astype(int)
    if "Mes" in df.columns:
        df["Mes"] = pd.to_numeric(df["Mes"], errors="coerce").fillna(0).astype(int)
    sem_opts = [{"label":f"Sem {s}","value":s} for s in sorted(df["Semana"].unique()) if s > 0] if "Semana" in df.columns else []
    mes_opts = [{"label":mes_map.get(m,str(m)),"value":m} for m in sorted(df["Mes"].unique()) if m > 0] if "Mes" in df.columns else []
    linea_opts = [{"label":l,"value":l} for l in sorted(df["Línea"].dropna().unique()) if l]
    sub_opts   = [{"label":s,"value":s} for s in sorted(df["Subcomponente"].dropna().unique()) if s]

    # Aplicar filtros
    df_fil = df.copy()
    if semanas and "Semana" in df_fil.columns: df_fil = df_fil[df_fil["Semana"].isin(semanas)]
    if mes_sel and "Mes" in df_fil.columns:    df_fil = df_fil[df_fil["Mes"] == int(mes_sel)]
    if lineas:                                  df_fil = df_fil[df_fil["Línea"].isin(lineas)]
    if subs:                                    df_fil = df_fil[df_fil["Subcomponente"].isin(subs)]

    if df_fil.empty:
        return [], dbc.Alert("Sin datos para este filtro.", color="info"), [], sem_opts, mes_opts, linea_opts, sub_opts

    # Totales del último mes disponible
    plan_total = df_fil["Plan semana"].sum() if "Plan semana" in df_fil.columns else 0
    real_total = df_fil["Real semana"].sum() if "Real semana" in df_fil.columns else 0
    pct_total  = round(real_total / plan_total * 100, 1) if plan_total > 0 else 0
    col_kpi    = "success" if pct_total >= 90 else "warning" if pct_total >= 70 else "danger"

    # KPIs
    ultimo_mes = plan_cols[-1].split("(")[0].strip() if plan_cols else ""
    kpis = dbc.Row([
        dbc.Col(make_kpi("% Cumplimiento", f"{pct_total}%", ultimo_mes, col_kpi), width=3),
        dbc.Col(make_kpi("Plan",  fmt_num(plan_total),  "unidades", "info"),    width=3),
        dbc.Col(make_kpi("Real",  fmt_num(real_total),  "unidades", "primary"), width=3),
        dbc.Col(make_kpi("Línea", str(df_fil["Línea"].nunique()), "en filtro"), width=3),
    ], className="g-2")

    # ── Gráfico 1: Plan vs Real por mes ──────────────────────────────────────
    df_mes_data = df.copy()
    if lineas: df_mes_data = df_mes_data[df_mes_data["Línea"].isin(lineas)]
    if subs:   df_mes_data = df_mes_data[df_mes_data["Subcomponente"].isin(subs)]
    if "Mes" in df_mes_data.columns:
        df_mes_data["Mes"] = pd.to_numeric(df_mes_data["Mes"], errors="coerce").fillna(0).astype(int)
        grp_mes = df_mes_data.groupby("Mes", as_index=False).agg(
            Plan=("Plan semana","sum"), Real=("Real semana","sum"))
        grp_mes["Mes_label"] = grp_mes["Mes"].map(mes_map).fillna(grp_mes["Mes"].astype(str))
        meses_labels = grp_mes["Mes_label"].tolist()
        plan_tots    = grp_mes["Plan"].tolist()
        real_tots    = grp_mes["Real"].tolist()
    else:
        meses_labels, plan_tots, real_tots = [], [], []
    pcts_mes = [round(r/p*100,1) if p>0 else 0 for r,p in zip(real_tots, plan_tots)]
    gaps     = [max(p-r,0) for p,r in zip(plan_tots, real_tots)]

    fig_mes = go.Figure()
    fig_mes.add_trace(go.Bar(
        name="Plan", x=meses_labels, y=plan_tots,
        marker_color="#1F3864",
        text=[f"{v:,.0f}" for v in plan_tots],
        textposition="inside", textfont=dict(size=10, color="white"),
    ))
    fig_mes.add_trace(go.Bar(
        name="Real", x=meses_labels, y=real_tots,
        marker_color="#2E75B6",
        text=[f"{v:,.0f} ({p}%)" for v,p in zip(real_tots,pcts_mes)],
        textposition="inside", textfont=dict(size=10, color="white"),
    ))
    fig_mes.add_trace(go.Bar(
        name="Gap", x=meses_labels, y=gaps,
        marker_color="#FF6B6B", opacity=0.7,
        base=real_tots,
        text=[f"▼{v:,.0f}" for v in gaps],
        textposition="outside", textfont=dict(size=9, color="#c00000"),
    ))
    fig_mes.update_layout(
        title=dict(text="Plan vs Real por Mes", font=dict(size=14, color="#1F3864")),
        barmode="overlay", plot_bgcolor="white", paper_bgcolor="white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=40,r=20,t=70,b=40), height=340,
        yaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0"),
    )

    # ── Gráfico 2: Plan vs Real por Subcomponente ────────────────────────────
    if "Plan semana" in df_fil.columns and "Real semana" in df_fil.columns:
        agg_d = {"Plan":("Plan semana","sum"), "Real":("Real semana","sum")}
        if "Planificado" in df_fil.columns:
            agg_d["Planif"] = ("Planificado","max")
        df_sub_grp = df_fil.groupby(["Línea","Subcomponente"], as_index=False).agg(**agg_d)
        if "Planif" not in df_sub_grp.columns:
            df_sub_grp["Planif"] = df_sub_grp["Plan"]
    df_sub_grp["Label"] = df_sub_grp["Línea"] + " — " + df_sub_grp["Subcomponente"]
    df_sub_grp = df_sub_grp.sort_values(["Línea","Plan"], ascending=[True,True])
    df_sub_grp["Pct"] = (df_sub_grp["Real"]/df_sub_grp["Plan"]*100).round(1).fillna(0)
    df_sub_grp["Gap"] = (df_sub_grp["Plan"] - df_sub_grp["Real"]).clip(lower=0)

    fig_sub = go.Figure()
    fig_sub.add_trace(go.Bar(
        name="Planificado", y=df_sub_grp["Label"], x=df_sub_grp["Planif"],
        orientation="h", marker_color="#D9E1F2", opacity=0.9,
    ))
    fig_sub.add_trace(go.Bar(
        name="Plan semana", y=df_sub_grp["Label"], x=df_sub_grp["Plan"],
        orientation="h", marker_color="#1F3864", opacity=0.7,
    ))
    fig_sub.add_trace(go.Bar(
        name="Real semana", y=df_sub_grp["Label"], x=df_sub_grp["Real"],
        orientation="h", marker_color="#2E75B6",
        text=[f"{p}%" for p in df_sub_grp["Pct"]],
        textposition="outside", textfont=dict(size=9),
    ))
    fig_sub.update_layout(
        title=dict(text="Plan vs Real por Subcomponente (período)", font=dict(size=14, color="#1F3864")),
        barmode="overlay", plot_bgcolor="white", paper_bgcolor="white",
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1),
        margin=dict(l=180,r=80,t=70,b=40),
        height=max(300, len(df_sub_grp)*30 + 120),
        xaxis=dict(tickformat=",.0f", gridcolor="#f0f0f0"),
    )

    # ── Gráfico 3: Dona % cumplimiento ────────────────────────────────────────
    pendiente = max(plan_total - real_total, 0)
    fig_dona = go.Figure(go.Pie(
        values=[real_total, pendiente],
        labels=["Real","Pendiente"],
        hole=0.65,
        marker_colors=["#2E75B6","#E9EFF7"],
        textinfo="none",
    ))
    fig_dona.add_annotation(
        text=f"<b>{pct_total}%</b><br><span style='font-size:11px'>Cumplimiento</span>",
        x=0.5, y=0.5, showarrow=False,
        font=dict(size=18, color="#1F3864"), align="center",
    )
    fig_dona.update_layout(
        title=dict(text="% Cumplimiento período", font=dict(size=14, color="#1F3864")),
        showlegend=False, margin=dict(l=20,r=20,t=60,b=20),
        height=300, paper_bgcolor="white",
    )

    graficos = html.Div([
        dbc.Row([
            dbc.Col(dcc.Graph(figure=fig_mes,  config={"displayModeBar":False}), width=8),
            dbc.Col(dcc.Graph(figure=fig_dona, config={"displayModeBar":False}), width=4),
        ], className="g-2 mb-3"),
        dbc.Row([
            dbc.Col(dcc.Graph(figure=fig_sub, config={"displayModeBar":False}), width=12),
        ], className="g-2"),
    ])

    # ── Tabla ─────────────────────────────────────────────────────────────────
    cols_show = ["Semana","Lunes","Viernes","O. Prod.","Máquina",
                 "Sec","Código","Descripción","Planificado","Cap. diaria",
                 "Plan semana","Real semana"]
    cols_show = [c for c in cols_show if c in df_fil.columns]
    num_c = ["Planificado","Cap. diaria","Plan semana","Real semana"]
    dt_cols = [{"name":c,"id":c,**({"type":"numeric","format":{"specifier":",.0f"}} if c in num_c else {})} for c in cols_show]

    style_cond = [
        {"if":{"filter_query":"{Plan semana} > 0 && {Real semana} >= {Plan semana}","column_id":"Real semana"},
         "backgroundColor":"#c6efce","color":"#1a7a4a"},
        {"if":{"filter_query":"{Plan semana} > 0 && {Real semana} < {Plan semana}","column_id":"Real semana"},
         "backgroundColor":"#ffc7ce","color":"#A32D2D"},
    ]

    tabla = dash_table.DataTable(
        columns=dt_cols,
        data=df_fil[cols_show].to_dict("records"),
        sort_action="native", page_size=25,
        style_table={"overflowX":"auto"},
        style_header={"backgroundColor":"#1F3864","color":"white","fontWeight":"500",
                      "fontSize":"10px","border":"0.5px solid #dee2e6","textAlign":"center"},
        style_cell={"fontSize":"10px","padding":"4px 7px",
                    "border":"0.5px solid #dee2e6","textAlign":"center"},
        style_data_conditional=style_cond,
        style_cell_conditional=[
            {"if":{"column_id":"Subcomponente"},"textAlign":"left"},
            {"if":{"column_id":"Línea"},"textAlign":"left"},
        ],
    )

    return kpis, graficos, tabla, sem_opts, mes_opts, linea_opts, sub_opts



# ── Callback: Última actualización ───────────────────────────────────────────

@app.callback(
    Output("last-update-time", "children"),
    Input("auto-refresh", "n_intervals"),
)
def actualizar_timestamp(_):
    from datetime import datetime as _dt
    return f"Actualizado: {_dt.now().strftime('%d/%m/%Y %H:%M')}"


# ── Main ──────────────────────────────────────────────────────────────────────

server = app.server  # para Render/gunicorn

if __name__ == "__main__":
    init_db(DB_PATH)
    if os.path.exists(EXCEL_PLAN):
        cargar_plan_desde_excel(DB_PATH, EXCEL_PLAN)
        cargar_bom_stock(DB_PATH, EXCEL_PLAN)
        print(f"✅ Plan, BOM y Stock cargados desde {EXCEL_PLAN}")
    else:
        print(f"⚠️  No se encontró {EXCEL_PLAN}")
    app.run(debug=True, port=8050)
