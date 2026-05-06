"""
exportar_excel.py — MPS_Reporte.xlsx
Por cada Proceso genera:
  - Hoja "Plan {Proceso}"  → tabla resumen
  - Hoja "Gantt {Proceso}" → columnas = días hábiles, filas = OPs
"""

import argparse
from datetime import date, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from db import init_db, cargar_plan_desde_excel, cargar_bom_stock, get_todas_ordenes, get_todos_ingresos

DB_PATH    = "mps_plasticos.db"
EXCEL_PLAN = "plan_produccion.xlsx"
SALIDA     = "MPS_Reporte.xlsx"

C_AZUL     = "1F3864"
C_TITULO   = "2F5496"
C_VERDE    = "C6EFCE"
C_AMARILLO = "FFEB9C"
C_ROJO     = "FFC7CE"
C_GRIS     = "EDEDED"
C_FUERA    = "F2F2F2"


def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def border_thin():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)


def dias_habiles_rango(desde: date, hasta: date):
    dias, d = [], desde
    while d <= hasta:
        if d.weekday() < 5:
            dias.append(d)
        d += timedelta(days=1)
    return dias


PROCESO_ALIAS = {
    "Masas y Tintas": "M&T",
    "Semiterminado":  "Semi",
    "Terminado":      "Terminado",
}

def nombre_hoja(prefix: str, proceso: str) -> str:
    """Nombre de hoja válido para Excel (max 31 chars)."""
    alias = PROCESO_ALIAS.get(proceso, proceso)
    nombre = f"{prefix} {alias}"
    return nombre[:31]


# ── Hoja Plan ─────────────────────────────────────────────────────────────────

def escribir_hoja_plan(ws, df: pd.DataFrame, proceso: str):
    encabezados = ["O. Prod.", "Máquina", "Línea", "Turnos", "Sec",
                   "Código", "Descripción", "Inicio", "Hora ini.",
                   "Fin", "Hora fin", "Dur.(hs)", "Planificado", "Cap. diaria"]
    anchos      = [11, 9, 12, 9, 5, 14, 40, 12, 10, 12, 10, 10, 14, 14]

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(encabezados))}1")
    ws["A1"].value     = f"Plan de Producción – {proceso}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill      = fill(C_TITULO)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c, (enc, ancho) in enumerate(zip(encabezados, anchos), 1):
        cell = ws.cell(2, c, enc)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = fill(C_AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border_thin()
        ws.column_dimensions[get_column_letter(c)].width = ancho
    ws.row_dimensions[2].height = 28

    colores, palette, idx = {}, ["EBF3FB", "FFFFFF"], 0
    for _, row in df.iterrows():
        if row["Maquina"] not in colores:
            colores[row["Maquina"]] = palette[idx % 2]
            idx += 1

    for fi, (_, row) in enumerate(df.iterrows(), 3):
        bg      = colores.get(row["Maquina"], "FFFFFF")
        turnos  = "2 turnos" if int(row["Horas_Prog"]) == 24 else "1 turno"
        vals = [
            int(row["BELNR_ID"]), row["Maquina"], row["Linea"], turnos, int(row["Sec"]),
            row["ItemCode"], row["Descripcion"],
            row["Fecha_Inicio"], row["Hora_Inicio"],
            row["Fecha_Fin"],    row["Hora_Fin"],
            round(float(row["Duracion_Horas"]), 1),
            int(row["Planificado"]), int(row["Cap_Diaria"]),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(fi, c, val)
            cell.fill      = fill(bg)
            cell.border    = border_thin()
            cell.alignment = Alignment(
                horizontal="left" if c == 7 else "center", vertical="center")
            cell.font = Font(size=9)
            if c in (13, 14):
                cell.number_format = "#,##0"

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(encabezados))}{len(df)+2}"


# ── Hoja Gantt ────────────────────────────────────────────────────────────────

def escribir_hoja_gantt(ws, df: pd.DataFrame, df_ingresos: pd.DataFrame,
                        proceso: str, desde: date, hasta: date, linea: str = None):
    from db import es_habil
    dias = [d for d in (desde + timedelta(days=i)
                        for i in range((hasta - desde).days + 1))
            if es_habil(d)]
    if not dias:
        ws["A1"] = "Sin días hábiles en el rango."
        return

    # Filtrar por línea si se especifica
    if linea:
        df = df[df["Linea"] == linea].reset_index(drop=True)

    # Índice ingresos por (belnr, fecha): cantidad acumulada real
    ingr_dia = {}
    if not df_ingresos.empty and "Cantidad_Real" in df_ingresos.columns:
        for _, r in df_ingresos.iterrows():
            ingr_dia[(int(r["BELNR_ID"]), str(r["Fecha"])[:10])] = float(r["Cantidad_Real"])

    COLS_FIJAS = 8  # OP/Máquina/Línea/Turnos/Sec/Código/Descripción/Planificado
    total_cols = COLS_FIJAS + len(dias)

    # Título
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(1, 1, f"Gantt Plan vs Real – {proceso} ({desde.strftime('%d/%m')} al {hasta.strftime('%d/%m/%Y')})")
    c.font = Font(bold=True, color="FFFFFF", size=12)
    c.fill = fill(C_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    fijos  = ["O. Prod.", "Máquina", "Línea", "Turnos", "Sec", "Código", "Descripción", "Planificado"]
    anchos = [11, 9, 12, 9, 5, 14, 38, 14]
    for c_idx, (enc, ancho) in enumerate(zip(fijos, anchos), 1):
        cell = ws.cell(2, c_idx, enc)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = fill(C_AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border_thin()
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws.row_dimensions[2].height = 36

    for d_idx, dia in enumerate(dias, COLS_FIJAS + 1):
        cell = ws.cell(2, d_idx, dia.strftime("%d/%m\n%a"))
        cell.font      = Font(bold=True, color="FFFFFF", size=8)
        cell.fill      = fill(C_AZUL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border_thin()
        ws.column_dimensions[get_column_letter(d_idx)].width = 16

    for fi, (_, row) in enumerate(df.iterrows(), 3):
        belnr    = int(row["BELNR_ID"])
        f_inicio = date.fromisoformat(str(row["Fecha_Inicio"])[:10])
        f_fin    = date.fromisoformat(str(row["Fecha_Fin"])[:10])
        cap_dia  = float(row["Cap_Diaria"])
        planif   = float(row["Planificado"])
        hp       = int(row["Horas_Prog"])
        velocidad = cap_dia / hp if hp > 0 else 0
        turnos   = "2 turnos" if hp == 24 else "1 turno"

        for c_idx, val in enumerate([
            belnr, row["Maquina"], row["Linea"], turnos, int(row["Sec"]),
            row["ItemCode"], row["Descripcion"], int(planif)
        ], 1):
            cell = ws.cell(fi, c_idx, val)
            cell.font      = Font(size=9)
            cell.alignment = Alignment(
                horizontal="left" if c_idx == 7 else "center", vertical="center")
            cell.border    = border_thin()
            if c_idx == 8:
                cell.number_format = "#,##0"

        # Calcular plan teórico acumulado y real acumulado día a día
        acum_plan_teo = 0.0  # plan teórico acumulado
        acum_real     = 0.0  # real acumulado

        for d_idx, dia in enumerate(dias, COLS_FIJAS + 1):
            cell = ws.cell(fi, d_idx)
            cell.border    = border_thin()
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.font      = Font(size=8)

            # Real del día (puede existir aunque sea fuera del rango teórico)
            real_dia = ingr_dia.get((belnr, dia.strftime("%Y-%m-%d")), None)

            if dia < f_inicio:
                cell.fill = fill(C_FUERA)
                continue

            if dia > f_fin:
                # Fuera del rango teórico — mostrar solo si hay real
                if real_dia is not None:
                    acum_real += real_dia
                    cell.value = "R:" + f"{int(real_dia):,}" + "\nA:" + f"{int(acum_real):,}"
                    cell.fill  = fill(C_AMARILLO)
                else:
                    cell.fill = fill(C_FUERA)
                continue

            # Plan del día teórico
            restante_plan = planif - acum_plan_teo
            plan_dia = min(cap_dia, max(restante_plan, 0))
            acum_plan_teo += plan_dia
            if real_dia is not None:
                acum_real += real_dia

            # "Debería ir" = plan teórico acumulado hasta este día
            deberia = acum_plan_teo

            if real_dia is None:
                cell.value = "P:" + f"{int(plan_dia):,}" + "\nD:" + f"{int(deberia):,}"
                cell.fill  = fill(C_GRIS)
            else:
                pct = real_dia / plan_dia * 100 if plan_dia > 0 else 0
                pct_acum = acum_real / deberia * 100 if deberia > 0 else 0
                cell.value = ("P:" + f"{int(plan_dia):,}" + "\n"
                              + "R:" + f"{int(real_dia):,}" + "\n"
                              + "A:" + f"{int(acum_real):,}" + "\n"
                              + "D:" + f"{int(deberia):,}" + "\n"
                              + f"{pct:.0f}%/{pct_acum:.0f}%")
                cell.fill  = fill(C_VERDE if pct >= 85 else C_AMARILLO if pct >= 65 else C_ROJO)

        ws.row_dimensions[fi].height = 62

    ws.freeze_panes = f"{get_column_letter(COLS_FIJAS + 1)}3"

    # Leyenda
    fila_ley = len(df) + 4
    ws.cell(fila_ley, 1, "Leyenda — P:Plan día  R:Real día  A:Acumulado real  D:Debería ir  %día/%acum").font = Font(bold=True, size=9)
    for i, (color, texto) in enumerate([
        (C_VERDE, "≥100%"), (C_AMARILLO, "75–99%"),
        (C_ROJO, "<75%"), (C_GRIS, "Sin registro"), (C_FUERA, "Fuera de rango")
    ]):
        ws.cell(fila_ley + 1, 2 + i*2).fill   = fill(color)
        ws.cell(fila_ley + 1, 2 + i*2).border = border_thin()
        ws.cell(fila_ley + 1, 3 + i*2, texto).font = Font(size=9)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--desde", default=None)
    parser.add_argument("--hasta", default=None)
    args = parser.parse_args()

    init_db(DB_PATH)
    cargar_plan_desde_excel(DB_PATH, EXCEL_PLAN)
    cargar_bom_stock(DB_PATH, EXCEL_PLAN)

    df_plan     = get_todas_ordenes(DB_PATH)
    df_ingresos = get_todos_ingresos(DB_PATH)

    if df_plan.empty:
        print("❌ No hay órdenes en el plan.")
        return

    fecha_min = date.fromisoformat(df_plan["Fecha_Inicio"].min()[:10])
    fecha_max = date.fromisoformat(df_plan["Fecha_Fin"].max()[:10])
    desde = date.fromisoformat(args.desde) if args.desde else fecha_min
    hasta = date.fromisoformat(args.hasta) if args.hasta else fecha_max

    print(f"📅 Rango: {desde} → {hasta}")

    wb = Workbook()
    wb.remove(wb.active)  # eliminar hoja vacía por defecto

    # Agrupar por Mes+Proceso para generar hojas separadas por mes
    meses = sorted(df_plan["Mes"].dropna().unique().astype(int)) if "Mes" in df_plan.columns else [0]

    procesos = sorted(df_plan["Proceso"].unique())
    for proceso in procesos:
        df_proc_full = df_plan[df_plan["Proceso"] == proceso].reset_index(drop=True)

        # Si hay columna Mes, generar una hoja por mes
        if "Mes" in df_plan.columns and df_plan["Mes"].notna().any():
            meses_proc = sorted(df_proc_full["Mes"].dropna().unique().astype(int))
            for mes in meses_proc:
                df_proc = df_proc_full[df_proc_full["Mes"].astype(int) == mes].reset_index(drop=True)
                if df_proc.empty:
                    continue
                alias_proc = PROCESO_ALIAS.get(proceso, proceso)
                suffix = f"M{mes}"
                print(f"  → {proceso} Mes {mes}: {len(df_proc)} órdenes")

                ws_plan = wb.create_sheet(title=f"Plan {alias_proc} {suffix}"[:31])
                escribir_hoja_plan(ws_plan, df_proc, f"{proceso} — Mes {mes}")

                # Rango de fechas del mes
                f_ini_mes = date.fromisoformat(df_proc["Fecha_Inicio"].min()[:10])
                f_fin_mes = date.fromisoformat(df_proc["Fecha_Fin"].max()[:10])
                ws_gantt = wb.create_sheet(title=f"Gantt {alias_proc} {suffix}"[:31])
                escribir_hoja_gantt(ws_gantt, df_proc, df_ingresos, f"{proceso} — Mes {mes}", f_ini_mes, f_fin_mes)
        else:
            print(f"  → {proceso}: {len(df_proc_full)} órdenes")
            ws_plan = wb.create_sheet(title=nombre_hoja("Plan", proceso))
            escribir_hoja_plan(ws_plan, df_proc_full, proceso)
            ws_gantt = wb.create_sheet(title=nombre_hoja("Gantt", proceso))
            escribir_hoja_gantt(ws_gantt, df_proc_full, df_ingresos, proceso, desde, hasta)

    wb.save(SALIDA)
    print(f"✅ Guardado: {SALIDA}")


if __name__ == "__main__":
    main()
