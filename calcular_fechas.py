"""
calcular_fechas.py — Calcula fechas de inicio/fin por secuencia de máquina
y exporta un Excel congelado que puede editarse manualmente.

Uso:
    python calcular_fechas.py
    python calcular_fechas.py --excel plan_produccion.xlsx --salida fechas_calculadas.xlsx

El Excel resultante (fechas_calculadas.xlsx) tiene una hoja "Fechas" con:
    BELNR_ID | Fecha_Inicio | Hora_Inicio | Fecha_Fin | Hora_Fin | Cap_Diaria | Duracion_Horas

Para mover fechas: editá directamente esa hoja y el app las tomará de ahí.
"""

import argparse
from datetime import date, datetime, timedelta

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Constantes ────────────────────────────────────────────────────────────────

EXCEL_PLAN = "plan_produccion.xlsx"
SALIDA     = "fechas_calculadas.xlsx"

HORA_INICIO  = 7   # 07:00
HORA_FIN_12  = 19  # 19:00 para máquinas de 1 turno
MAQUINAS_PARALELAS = {"AMAN02"}

# Feriados — se cargan desde la hoja Feriados del Excel
FERIADOS = set()

# Mantenimientos por máquina
MANTENIMIENTOS = {}


# ── Helpers de calendario ─────────────────────────────────────────────────────

def es_habil(d: date, maquina: str = None) -> bool:
    if d.weekday() >= 5 or d in FERIADOS:
        return False
    if maquina and d in MANTENIMIENTOS.get(maquina, set()):
        return False
    return True


def siguiente_momento_habil(dt: datetime, horas_prog: int, maquina: str = None) -> datetime:
    MAX_ITER = 60
    for _ in range(MAX_ITER):
        d_date = dt.date()
        mant_fechas = MANTENIMIENTOS.get(maquina, set()) if maquina else set()
        if d_date in FERIADOS or d_date in mant_fechas:
            dt = datetime(dt.year, dt.month, dt.day, HORA_INICIO) + timedelta(days=1)
            continue
        if horas_prog == 24:
            if dt.weekday() == 5 and dt.hour >= HORA_INICIO:
                dt = datetime(dt.year, dt.month, dt.day, HORA_INICIO) + timedelta(days=2)
                continue
            if dt.weekday() == 6:
                dt = datetime(dt.year, dt.month, dt.day, HORA_INICIO) + timedelta(days=1)
                continue
        else:
            if dt.weekday() < 5 and dt.hour >= HORA_FIN_12:
                dt = datetime(dt.year, dt.month, dt.day, HORA_INICIO) + timedelta(days=1)
                continue
            if dt.weekday() >= 5:
                days_fwd = (7 - dt.weekday()) % 7 or 7
                dt = datetime(dt.year, dt.month, dt.day, HORA_INICIO) + timedelta(days=days_fwd)
                continue
            if dt.hour < HORA_INICIO:
                dt = datetime(dt.year, dt.month, dt.day, HORA_INICIO)
        break
    return dt


def sumar_horas_habiles(inicio: datetime, horas: float,
                         horas_prog: int, maquina: str = None) -> datetime:
    restante = horas
    actual   = siguiente_momento_habil(inicio, horas_prog, maquina)

    while restante > 0:
        if horas_prog == 24:
            # Fin del día hábil actual = 07:00 del próximo día hábil
            corte_raw = datetime(actual.year, actual.month, actual.day, HORA_INICIO) + timedelta(days=1)
            corte = siguiente_momento_habil(corte_raw, horas_prog, maquina)
            # Horas disponibles hoy = máximo 24hs (un turno)
            horas_disp = min((corte - actual).total_seconds() / 3600, 24.0)
            if restante <= horas_disp:
                actual = actual + timedelta(hours=restante)
                restante = 0
            else:
                restante -= horas_disp
                actual = corte
        else:
            corte = datetime(actual.year, actual.month, actual.day, HORA_FIN_12)
            horas_disp = (corte - actual).total_seconds() / 3600
            if restante <= horas_disp:
                actual = actual + timedelta(hours=restante)
                restante = 0
            else:
                restante -= horas_disp
                siguiente = datetime(actual.year, actual.month, actual.day,
                                     HORA_INICIO) + timedelta(days=1)
                actual = siguiente_momento_habil(siguiente, horas_prog, maquina)

    return actual


def limpiar_cantidad(serie):
    return pd.to_numeric(serie, errors="coerce").fillna(0)


# ── Cargar feriados y mantenimientos ─────────────────────────────────────────

def cargar_calendario(excel_path: str):
    global FERIADOS, MANTENIMIENTOS
    try:
        df_fer = pd.read_excel(excel_path, sheet_name="Feriados")
        for _, r in df_fer.iterrows():
            try:
                f = pd.to_datetime(r.iloc[0], dayfirst=True).date()
                FERIADOS.add(f)
            except Exception:
                pass
        print(f"  → {len(FERIADOS)} feriados cargados")
    except Exception as e:
        print(f"  ⚠️  Feriados no cargados: {e}")

    try:
        df_mant = pd.read_excel(excel_path, sheet_name="Mantenimientos")
        df_mant.columns = df_mant.columns.str.strip()
        mant_dict = {}
        for _, r in df_mant.iterrows():
            try:
                maq   = str(r["Maquina"]).strip()
                f_ini = pd.to_datetime(r["Fecha_Inicio"], dayfirst=True).date()
                f_fin = pd.to_datetime(r["Fecha_Fin"],    dayfirst=True).date()
                if maq not in mant_dict:
                    mant_dict[maq] = set()
                d = f_ini
                while d <= f_fin:
                    mant_dict[maq].add(d)
                    d += timedelta(days=1)
            except Exception:
                pass
        MANTENIMIENTOS = mant_dict
        print(f"  → {sum(len(v) for v in mant_dict.values())} días de mantenimiento cargados")
    except Exception as e:
        print(f"  ⚠️  Mantenimientos no cargados: {e}")


# ── Calcular fechas ───────────────────────────────────────────────────────────

def calcular_fechas(excel_path: str) -> pd.DataFrame:
    # Leer Excel
    try:
        df = pd.read_excel(excel_path, sheet_name="Ordenes", dtype={"ItemCode": str})
    except Exception:
        df = pd.read_excel(excel_path, sheet_name=0, dtype={"ItemCode": str})

    df.columns = df.columns.str.strip()
    df = df.rename(columns={
        "Descripción del artículo": "Descripcion",
        "Descripción":              "Descripcion",
    })
    df = df.dropna(subset=["BELNR_ID"])
    df["BELNR_ID"] = df["BELNR_ID"].astype(int)

    # Limpiar columnas
    df["Cantidad_Base"]  = limpiar_cantidad(df.get("Cantidad Base", df.get("Cantidad_Base", 0)))
    df["Horas"]          = limpiar_cantidad(df.get("Horas", 0))
    df["Planificado"]    = limpiar_cantidad(df["Planificado"])
    df["Horas_Prog"]     = limpiar_cantidad(df.get("Horas Programadas", df.get("Horas_Prog", 12))).astype(int)
    df["Mes"]            = limpiar_cantidad(df.get("Mes", 0)).astype(int)
    df["Fecha_Inicio_Base"] = pd.to_datetime(df["Fecha Inicio"], dayfirst=True).dt.normalize()

    # Columnas opcionales
    for col in ["Estado", "Avance_SAP", "Fecha_Conta", "Fecha_Gantt_Desde",
                "Subcomponente", "Línea"]:
        if col not in df.columns:
            df[col] = ""

    df["Estado_OP"]  = df["Estado"].fillna("ABIERTO").str.strip().str.upper()
    df["Maquina"]    = df["Maquina"].str.strip() if "Maquina" in df.columns else df[" Maquina "].str.strip()

    registros = []

    # Agrupar por Mes + Máquina — cursor reinicia por cada combinación
    for (mes, maquina), grupo in df.groupby(["Mes", "Maquina"], sort=True):
        grupo = grupo.sort_values("Sec").reset_index(drop=True)
        horas_prog  = int(grupo.loc[0, "Horas_Prog"])
        es_paralela = maquina in MAQUINAS_PARALELAS

        primera_fecha = grupo.loc[0, "Fecha_Inicio_Base"]
        cursor = datetime(int(primera_fecha.year), int(primera_fecha.month),
                          int(primera_fecha.day), HORA_INICIO)
        cursor = siguiente_momento_habil(cursor, horas_prog, maquina)

        for _, row in grupo.iterrows():
            hp        = int(row["Horas_Prog"])
            horas     = float(row["Horas"]) if float(row["Horas"]) > 0 else 1
            velocidad = float(row["Cantidad_Base"]) / horas
            cap_dia   = velocidad * hp
            duracion  = float(row["Planificado"]) / velocidad if velocidad > 0 else 0

            if es_paralela:
                f_base    = row["Fecha_Inicio_Base"]
                dt_inicio = siguiente_momento_habil(
                    datetime(f_base.year, f_base.month, f_base.day, HORA_INICIO),
                    hp, maquina)
                dt_fin = sumar_horas_habiles(dt_inicio, duracion, hp, maquina)
            else:
                # Si la fecha del Excel es posterior al cursor, respetar la del Excel
                f_base = row["Fecha_Inicio_Base"]
                fecha_excel = siguiente_momento_habil(
                    datetime(f_base.year, f_base.month, f_base.day, HORA_INICIO),
                    hp, maquina)
                dt_inicio = max(cursor, fecha_excel)
                dt_fin    = sumar_horas_habiles(dt_inicio, duracion, hp, maquina)
                cursor    = dt_fin

            registros.append({
                "BELNR_ID":       int(row["BELNR_ID"]),
                "Mes":            mes,
                "Maquina":        maquina,
                "ItemCode":       str(row["ItemCode"]).strip(),
                "Descripcion":    row["Descripcion"],
                "Sec":            int(row["Sec"]),
                "Proceso":        row["Proceso"],
                "Linea":          str(row.get("Línea", row.get("Linea",""))).strip(),
                "Subcomponente":  str(row.get("Subcomponente","")).strip(),
                "Horas_Prog":     hp,
                "Planificado":    int(row["Planificado"]),
                "Cap_Diaria":     round(cap_dia),
                "Duracion_Horas": round(duracion, 2),
                "Fecha_Inicio":   dt_inicio.strftime("%Y-%m-%d"),
                "Hora_Inicio":    dt_inicio.strftime("%H:%M"),
                "Fecha_Fin":      dt_fin.strftime("%Y-%m-%d"),
                "Hora_Fin":       dt_fin.strftime("%H:%M"),
                "Estado_OP":      row["Estado_OP"],
                "Avance_SAP":     float(limpiar_cantidad(pd.Series([row.get("Avance_SAP", 0)])).iloc[0]),
                "Fecha_Conta":    str(pd.to_datetime(row["Fecha_Conta"], dayfirst=True).date())
                                  if pd.notna(row.get("Fecha_Conta")) and str(row.get("Fecha_Conta","")).strip()
                                  else dt_inicio.strftime("%Y-%m-%d"),
                "Fecha_Gantt_Desde": str(pd.to_datetime(row["Fecha_Gantt_Desde"], dayfirst=True).date())
                                     if pd.notna(row.get("Fecha_Gantt_Desde")) and str(row.get("Fecha_Gantt_Desde","")).strip()
                                     else dt_inicio.strftime("%Y-%m-%d"),
            })

    df_result = pd.DataFrame(registros)
    return df_result


# ── Exportar Excel ────────────────────────────────────────────────────────────

def exportar_excel(df: pd.DataFrame, salida: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Fechas"

    s = Side(style="thin", color="BFBFBF")
    brd = Border(left=s, right=s, top=s, bottom=s)

    # Columnas a exportar
    cols = [
        ("BELNR_ID",         "O. Prod.",        9),
        ("Mes",              "Mes",              5),
        ("Maquina",          "Máquina",          9),
        ("Proceso",          "Proceso",          14),
        ("Linea",            "Línea",            12),
        ("Subcomponente",    "Subcomponente",    14),
        ("Sec",              "Sec",              5),
        ("ItemCode",         "Código",           14),
        ("Descripcion",      "Descripción",      40),
        ("Horas_Prog",       "Hs. Prog.",        8),
        ("Planificado",      "Planificado",      13),
        ("Cap_Diaria",       "Cap. diaria",      13),
        ("Duracion_Horas",   "Dur.(hs)",         9),
        ("Fecha_Inicio",     "Fecha Inicio",     13),
        ("Hora_Inicio",      "Hora Ini.",        10),
        ("Fecha_Fin",        "Fecha Fin",        13),
        ("Hora_Fin",         "Hora Fin",         10),
        ("Estado_OP",        "Estado",           10),
        ("Avance_SAP",       "Avance SAP",       13),
        ("Fecha_Conta",      "Fecha Conta",      13),
        ("Fecha_Gantt_Desde","Fecha Gantt",      13),
    ]

    # Título
    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    ws["A1"] = "Fechas Calculadas — MPS Layconsa"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill      = PatternFill("solid", fgColor="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    # Encabezados
    for c_idx, (col_id, col_name, ancho) in enumerate(cols, 1):
        cell = ws.cell(2, c_idx, col_name)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = brd
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws.row_dimensions[2].height = 28

    # Colores alternados por máquina
    colores = {}
    palette = ["EBF3FB", "FFFFFF"]
    idx = 0
    for maq in df["Maquina"].unique():
        colores[maq] = palette[idx % 2]
        idx += 1

    # Datos
    for fi, (_, row) in enumerate(df.iterrows(), 3):
        bg = colores.get(row["Maquina"], "FFFFFF")
        for c_idx, (col_id, _, _) in enumerate(cols, 1):
            val = row.get(col_id, "")
            cell = ws.cell(fi, c_idx, val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = brd
            cell.font      = Font(size=9)
            cell.alignment = Alignment(
                horizontal="left" if c_idx == 9 else "center",
                vertical="center")
            if col_id in ("Planificado","Cap_Diaria","Avance_SAP"):
                cell.number_format = "#,##0"
        ws.row_dimensions[fi].height = 15

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(df)+2}"

    # ── Hoja Para Producción ──────────────────────────────────────────────────
    ws2 = wb.create_sheet(title="Para Producción")

    cols2 = [
        ("BELNR_ID",       "O. Prod.",    9),
        ("Mes",            "Mes",         5),
        ("Maquina",        "Máquina",     9),
        ("Sec",            "Sec",         5),
        ("ItemCode",       "Código",      14),
        ("Descripcion",    "Descripción", 42),
        ("Horas_Prog",     "Hs. Prog.",   8),
        ("Planificado",    "Planificado", 13),
        ("Cap_Diaria",     "Cap. diaria", 13),
        ("Duracion_Horas", "Dur.(hs)",    9),
        ("Fecha_Inicio",   "Fecha Inicio",13),
        ("Fecha_Fin",      "Fecha Fin",   13),
        ("Avance_SAP",     "Avance SAP",  13),
    ]

    # Título
    ws2.merge_cells(f"A1:{get_column_letter(len(cols2))}1")
    ws2["A1"] = "Plan de Producción — Entrega a Producción"
    ws2["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws2["A1"].fill      = PatternFill("solid", fgColor="2F5496")
    ws2["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 22

    for c_idx, (_, col_name, ancho) in enumerate(cols2, 1):
        cell = ws2.cell(2, c_idx, col_name)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = brd
        ws2.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws2.row_dimensions[2].height = 28

    # Colores alternados por máquina
    colores2 = {}
    palette2 = ["EBF3FB", "FFFFFF"]
    idx2 = 0
    for maq in df["Maquina"].unique():
        colores2[maq] = palette2[idx2 % 2]
        idx2 += 1

    for fi, (_, row) in enumerate(df.iterrows(), 3):
        bg = colores2.get(row["Maquina"], "FFFFFF")
        for c_idx, (col_id, _, _) in enumerate(cols2, 1):
            val = row.get(col_id, "")
            cell = ws2.cell(fi, c_idx, val)
            cell.fill      = PatternFill("solid", fgColor=bg)
            cell.border    = brd
            cell.font      = Font(size=9)
            cell.alignment = Alignment(
                horizontal="left" if c_idx == 5 else "center",
                vertical="center")
            if col_id in ("Planificado","Cap_Diaria","Avance_SAP"):
                cell.number_format = "#,##0"
        ws2.row_dimensions[fi].height = 15

    ws2.freeze_panes = "A3"
    ws2.auto_filter.ref = f"A2:{get_column_letter(len(cols2))}{len(df)+2}"

    # ── Hoja Mantenimientos ───────────────────────────────────────────────────
    try:
        df_mant = pd.read_excel(EXCEL_PLAN, sheet_name="Mantenimientos")
        df_mant.columns = df_mant.columns.str.strip()
        if not df_mant.empty:
            ws3 = wb.create_sheet(title="Mantenimientos")
            mant_cols = ["Maquina","Fecha_Inicio","Fecha_Fin","Descripcion"]
            mant_cols = [c for c in mant_cols if c in df_mant.columns]
            # Header
            for ci, cn in enumerate(mant_cols, 1):
                cell = ws3.cell(1, ci, cn)
                cell.font      = Font(bold=True, color="FFFFFF", size=9)
                cell.fill      = PatternFill("solid", fgColor="1F3864")
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border    = brd
            # Data
            for fi, (_, mrow) in enumerate(df_mant[mant_cols].iterrows(), 2):
                for ci, val in enumerate(mrow.values, 1):
                    cell = ws3.cell(fi, ci, val)
                    cell.font      = Font(size=9)
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                    cell.border    = brd
            ws3.column_dimensions["A"].width = 12
            ws3.column_dimensions["B"].width = 14
            ws3.column_dimensions["C"].width = 14
            ws3.column_dimensions["D"].width = 30
            print(f"  → {len(df_mant)} mantenimientos agregados")
    except Exception as e:
        print(f"  ⚠️  Mantenimientos no incluidos: {e}")

    wb.save(salida)
    print(f"✅ Exportado: {salida} ({len(df)} órdenes)")
    print()
    print("📌 Para ajustar fechas manualmente:")
    print("   1. Abrí fechas_calculadas.xlsx")
    print("   2. Editá las columnas Fecha_Inicio / Hora_Inicio / Fecha_Fin / Hora_Fin")
    print("   3. Guardá el archivo")
    print("   4. Reiniciá el app — leerá las fechas de este Excel")


# ── Main ──────────────────────────────────────────────────────────────────────


def exportar_cumplimiento_semanal(df: pd.DataFrame, semana: int):
    """
    Genera Excel con cumplimiento semanal por OP.
    semana: número de semana ISO (ej: 19 = semana del 05/05 al 09/05/2026)
    """
    import sqlite3 as _sq3

    año = date.today().year
    lunes  = date.fromisocalendar(año, semana, 1)
    viernes = date.fromisocalendar(año, semana, 5)
    print(f"  → Semana {semana}: {lunes.strftime('%d/%m/%Y')} – {viernes.strftime('%d/%m/%Y')}")

    # OPs activas en la semana
    df_sem = df[
        (df["Fecha_Inicio"] <= str(viernes)) &
        (df["Fecha_Fin"]    >= str(lunes))
    ].reset_index(drop=True)

    if df_sem.empty:
        print("  ⚠️  Sin órdenes para esta semana")
        return

    # Avance real de la semana
    try:
        con = _sq3.connect(DB_PATH)
        df_av = pd.read_sql(
            "SELECT BELNR_ID, SUM(Cantidad_Real) as real_sem FROM avance_real "
            "WHERE Fecha >= ? AND Fecha <= ? GROUP BY BELNR_ID",
            con, params=[str(lunes), str(viernes)])
        con.close()
        av_map = dict(zip(df_av["BELNR_ID"].astype(int), df_av["real_sem"].fillna(0)))
    except Exception:
        av_map = {}

    rows = []
    for _, row in df_sem.iterrows():
        f_ini_op = date.fromisoformat(str(row["Fecha_Inicio"])[:10])
        f_fin_op = date.fromisoformat(str(row["Fecha_Fin"])[:10])
        cap      = float(row["Cap_Diaria"])
        planif   = float(row["Planificado"])

        inicio_sem = max(f_ini_op, lunes)
        fin_sem    = min(f_fin_op, viernes)

        # Plan acumulado hasta inicio de semana
        acum_prev = 0.0
        d = f_ini_op
        while d < inicio_sem:
            if es_habil(d):
                acum_prev += min(cap, max(planif - acum_prev, 0))
            d += timedelta(days=1)

        # Plan de la semana
        plan_sem = 0.0
        d = inicio_sem
        while d <= fin_sem:
            if es_habil(d):
                dia_plan = min(cap, max(planif - acum_prev, 0))
                plan_sem  += dia_plan
                acum_prev += dia_plan
            d += timedelta(days=1)

        real_sem = float(av_map.get(int(row["BELNR_ID"]), 0))
        pct      = round(real_sem / plan_sem * 100, 1) if plan_sem > 0 else 0

        rows.append({
            "O. Prod.":    int(row["BELNR_ID"]),
            "Mes":         int(row["Mes"]),
            "Máquina":     row["Maquina"],
            "Proceso":     row["Proceso"],
            "Sec":         int(row["Sec"]),
            "Código":      row["ItemCode"],
            "Descripción": row["Descripcion"],
            "Plan semana": round(plan_sem),
            "Real semana": round(real_sem),
            "% Cumpl.":    pct,
        })

    df_out = pd.DataFrame(rows)

    # Exportar Excel
    wb = Workbook()
    ws = wb.active
    ws.title = f"Semana {semana}"

    s = Side(style="thin", color="BFBFBF")
    brd = Border(left=s, right=s, top=s, bottom=s)

    cols = [
        ("O. Prod.",    9), ("Mes",     5), ("Máquina",  9),
        ("Proceso",    14), ("Sec",     5), ("Código",  14),
        ("Descripción",42), ("Plan semana",14), ("Real semana",14), ("% Cumpl.",10),
    ]

    ws.merge_cells(f"A1:{get_column_letter(len(cols))}1")
    ws["A1"] = f"Cumplimiento Semana {semana} — {lunes.strftime('%d/%m')} al {viernes.strftime('%d/%m/%Y')}"
    ws["A1"].font      = Font(bold=True, color="FFFFFF", size=12)
    ws["A1"].fill      = PatternFill("solid", fgColor="2F5496")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22

    for c_idx, (col_name, ancho) in enumerate(cols, 1):
        cell = ws.cell(2, c_idx, col_name)
        cell.font      = Font(bold=True, color="FFFFFF", size=9)
        cell.fill      = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = brd
        ws.column_dimensions[get_column_letter(c_idx)].width = ancho
    ws.row_dimensions[2].height = 28

    for fi, (_, row) in enumerate(df_out.iterrows(), 3):
        vals = [row[c[0]] for c in cols]
        for c_idx, val in enumerate(vals, 1):
            cell = ws.cell(fi, c_idx, val)
            cell.border    = brd
            cell.font      = Font(size=9)
            cell.alignment = Alignment(
                horizontal="left" if c_idx == 7 else "center",
                vertical="center")
            if c_idx in (8, 9):
                cell.number_format = "#,##0"
            if c_idx == 10:
                cell.number_format = "0.0"
                pct_val = float(val) if val else 0
                if   pct_val >= 90: cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif pct_val >= 70: cell.fill = PatternFill("solid", fgColor="FFEB9C")
                else:               cell.fill = PatternFill("solid", fgColor="FFC7CE")
        ws.row_dimensions[fi].height = 15

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{get_column_letter(len(cols))}{len(df_out)+2}"

    salida = f"Cumplimiento_Semana{semana}_{lunes.strftime('%d%m')}.xlsx"
    wb.save(salida)
    print(f"  ✅ Guardado: {salida} ({len(df_out)} órdenes)")


def main():
    parser = argparse.ArgumentParser(
        description="Calcula fechas de producción y exporta Excel congelado")
    parser.add_argument("--excel",  default=EXCEL_PLAN, help="Plan de producción Excel")
    parser.add_argument("--salida", default=SALIDA,     help="Archivo de salida")
    parser.add_argument("--semana", type=int, default=None,
                        help="Número de semana ISO para cumplimiento (ej: 19)")
    args = parser.parse_args()

    print(f"📂 Leyendo: {args.excel}")
    cargar_calendario(args.excel)

    print("⚙️  Calculando fechas...")
    df = calcular_fechas(args.excel)

    if df.empty:
        print("❌ No se encontraron órdenes.")
        return

    # Resumen por máquina
    print(f"\n📅 Resumen ({len(df)} órdenes):")
    for maq, grp in df.groupby("Maquina"):
        f_ini = grp["Fecha_Inicio"].min()
        f_fin = grp["Fecha_Fin"].max()
        print(f"   {maq:12s}: {len(grp):3d} OPs  {f_ini} → {f_fin}")

    exportar_excel(df, args.salida)

    if args.semana:
        exportar_cumplimiento_semanal(df, args.semana)


if __name__ == "__main__":
    main()
